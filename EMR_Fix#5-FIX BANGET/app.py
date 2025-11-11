import os
import re
import base64
import io
import pyotp # Untuk 2FA
import qrcode # Untuk 2FA
import uuid
import hashlib
import secrets
from flask_mail import Mail, Message
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from itsdangerous import URLSafeTimedSerializer
from threading import Thread
from datetime import datetime, timedelta, timezone
from cryptography.fernet import Fernet
from functools import wraps
from io import BytesIO
from apscheduler.schedulers.background import BackgroundScheduler
from sqlalchemy import func, extract
from collections import defaultdict
from pathlib import Path

basedir = os.path.abspath(os.path.dirname(__file__))

def validate_image_file(file, max_size_mb=5):
    """
    Validasi file gambar yang diupload
    Returns: (is_valid: bool, error_message: str)
    """
    if not file or not file.filename:
        return False, "File tidak ditemukan"
    
    # Cek ekstensi file
    if not allowed_file(file.filename):
        return False, f"Format file tidak didukung. Gunakan: {', '.join(ALLOWED_EXTENSIONS)}"
    
    # Cek ukuran file (reset pointer dulu)
    file.seek(0, 2)  # Pindah ke akhir file
    file_size = file.tell()  # Dapatkan ukuran
    file.seek(0)  # Reset pointer ke awal
    
    max_size_bytes = max_size_mb * 1024 * 1024
    if file_size > max_size_bytes:
        return False, f"Ukuran file terlalu besar. Maksimal {max_size_mb}MB"
    
    return True, None

def load_key():
    try:
        return open("key.key", "rb").read()
    except FileNotFoundError:
        print("="*60)
        print("FATAL ERROR: File 'key.key' tidak ditemukan!")
        print("Jalankan script 'generate_key.py' terlebih dahulu untuk membuat kunci.")
        print("="*60)
        exit()

key = load_key()
fernet = Fernet(key)

UPLOAD_FOLDER = os.path.join(basedir, 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
MAX_FILE_SIZE = 5 * 1024 * 1024

app = Flask(__name__)
app.config['SECRET_KEY'] = 'kunci-rahasia-yang-super-aman-dan-unik'
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@localhost/emr_project_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_SERVER'] = 'mail.erkaes.my.id'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USERNAME'] = 'noreply@erkaes.my.id'
app.config['MAIL_PASSWORD'] = 'PapaBuangLimbah#312'
app.config['MAIL_DEFAULT_SENDER'] = ('Batam Sehat', 'noreply@erkaes.my.id')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["500 per day", "100 per hour"],
    storage_uri="memory://",
    strategy="fixed-window"
)

def reset_rate_limiter_for_user():
    """Reset rate limiter untuk IP user yang berhasil login"""
    try:
        # Hapus rate limit untuk IP ini
        limiter.reset()
        print(f"✅ Rate limiter direset untuk IP: {request.remote_addr}")
    except Exception as e:
        print(f"⚠️ Gagal reset rate limiter: {e}")

os.makedirs(os.path.join(UPLOAD_FOLDER, 'profiles'), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'ktp'), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'documents'), exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)
mail = Mail(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = "Silakan login untuk mengakses halaman ini."
login_manager.login_message_category = "warning"


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(50), nullable=False)
    
    email = db.Column(db.String(150), unique=True, nullable=True)
    nik_encrypted = db.Column(db.LargeBinary(512), nullable=True)
    full_name_encrypted = db.Column(db.LargeBinary(1024), nullable=True)
    contact_encrypted = db.Column(db.LargeBinary(512), nullable=True)
    date_of_birth_encrypted = db.Column(db.LargeBinary(256), nullable=True)
    address_encrypted = db.Column(db.LargeBinary(1024), nullable=True)
    
    verification_status = db.Column(db.String(50), nullable=True, default='belum diverifikasi')
    
    # Kolom untuk 2FA
    otp_secret = db.Column(db.String(32), nullable=True)
    
    is_active_db = db.Column(db.Boolean, default=True, nullable=False)
    profile_photo = db.Column(db.String(255), nullable=True)
    ktp_photo = db.Column(db.String(255), nullable=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def get_reset_token(self, expires_sec=1800):
        """Membuat token reset password (valid 30 menit)."""
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        return s.dumps(self.id, salt='password-reset-salt')

    @staticmethod
    def verify_reset_token(token, expires_sec=1800):
        """Memverifikasi token dan mengembalikan user jika valid."""
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        try:
            user_id = s.loads(
                token,
                salt='password-reset-salt',
                max_age=expires_sec
            )
        except:
            return None # Token tidak valid atau kedaluwarsa
        return User.query.get(user_id)
    
    def get_email_change_token(self, new_email, expires_sec=1800):
        """Membuat token konfirmasi ganti email (valid 30 menit)."""
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        # Simpan ID user dan email BARU dalam token
        return s.dumps({'user_id': self.id, 'new_email': new_email}, salt='email-confirm-salt')

    @staticmethod
    def verify_email_change_token(token, expires_sec=1800):
        """Memverifikasi token ganti email & mengembalikan data jika valid."""
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        try:
            data = s.loads(
                token,
                salt='email-confirm-salt',
                max_age=expires_sec
            )
            # Pastikan format data benar
            if 'user_id' in data and 'new_email' in data:
                return data
        except:
            return None # Token tidak valid, kedaluwarsa, atau format salah
        return None

    @property
    def nik(self):
        if not self.nik_encrypted: return "Belum diisi"
        try: return fernet.decrypt(self.nik_encrypted).decode()
        except: return "N/A"

    @property
    def full_name(self):
        if not self.full_name_encrypted: return self.username
        try: return fernet.decrypt(self.full_name_encrypted).decode()
        except: return self.username

    @property
    def contact(self):
        if not self.contact_encrypted: return "Belum diisi"
        try: return fernet.decrypt(self.contact_encrypted).decode()
        except: return "N/A"

    @property
    def date_of_birth(self):
        if not self.date_of_birth_encrypted: return None
        try:
            date_str = fernet.decrypt(self.date_of_birth_encrypted).decode()
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except: return None
    
    @property
    def address(self):
        if not self.address_encrypted: return "Belum diisi"
        try: return fernet.decrypt(self.address_encrypted).decode()
        except: return "N/A"
    
    @property
    def is_profile_complete(self):
        return self.nik_encrypted and self.full_name_encrypted and self.contact_encrypted and self.date_of_birth_encrypted and self.address_encrypted
    
    @property
    def is_active(self):
        return True

class Patient(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    medical_record_number = db.Column(db.String(20), unique=True, nullable=False)
    full_name_encrypted = db.Column(db.LargeBinary(1024), nullable=False)
    date_of_birth_encrypted = db.Column(db.LargeBinary(256), nullable=False)
    address_encrypted = db.Column(db.LargeBinary(1024), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=True)
    user = db.relationship('User', backref=db.backref('patient_record', uselist=False))
    profile_photo = db.Column(db.String(255), nullable=True)  # Path foto profil
    ktp_photo = db.Column(db.String(255), nullable=True)
    visits = db.relationship('Visit', backref='patient', lazy=True, cascade="all, delete-orphan")
    documents = db.relationship('PatientDocument', backref='patient', lazy=True, cascade="all, delete-orphan")  # Relasi baru

    @property
    def full_name(self):
        # Ambil data dari User (profil) jika terhubung
        if self.user:
            return self.user.full_name
        # Fallback jika user tidak terhubung (misal: data lama)
        try: return fernet.decrypt(self.full_name_encrypted).decode()
        except: return "Error Dekripsi"
            
    @property
    def date_of_birth(self):
        # Ambil data dari User (profil) jika terhubung
        if self.user:
            return self.user.date_of_birth
        # Fallback jika user tidak terhubung
        try:
            date_str = fernet.decrypt(self.date_of_birth_encrypted).decode()
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except: return None

    @property
    def address(self):
        # Ambil data dari User (profil) jika terhubung
        if self.user:
            return self.user.address
        # Fallback jika user tidak terhubung
        try: return fernet.decrypt(self.address_encrypted).decode()
        except: return "Error Dekripsi"
        
    @property
    def age(self):
        if not self.date_of_birth:
            return None
        today = datetime.today().date()
        return today.year - self.date_of_birth.year - ((today.month, today.day) < (self.date_of_birth.month, self.date_of_birth.day))

class PatientDocument(db.Model):
    """Model untuk menyimpan dokumen medis pasien (hasil lab, X-Ray, dll)"""
    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('patient.id'), nullable=False)
    document_type = db.Column(db.String(50), nullable=False) 
    file_name = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_size = db.Column(db.Integer, nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    upload_date = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    description = db.Column(db.Text, nullable=True)
    
    uploader = db.relationship('User', backref='uploaded_documents')

class Visit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    visit_date = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    patient_id = db.Column(db.Integer, db.ForeignKey('patient.id'), nullable=False)
    doctor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    appointment_id = db.Column(db.Integer, db.ForeignKey('appointment.id'), nullable=True, unique=True)
    appointment = db.relationship('Appointment', backref=db.backref('visit', uselist=False))
    doctor = db.relationship('User', backref='visits')
    soap_note = db.relationship('SoapNote', backref='visit', uselist=False, cascade="all, delete-orphan")

class SoapNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subjective = db.Column(db.Text, nullable=True)
    objective = db.Column(db.Text, nullable=True)
    assessment = db.Column(db.Text, nullable=True)
    plan = db.Column(db.Text, nullable=True)
    visit_id = db.Column(db.Integer, db.ForeignKey('visit.id'), nullable=False)

class AppointmentSlot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    doctor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    start_time = db.Column(db.DateTime, nullable=False)
    end_time = db.Column(db.DateTime, nullable=False)
    status = db.Column(db.String(50), nullable=False, default='Available') 
    verification_status = db.Column(db.String(50), nullable=False, default='Pending')
    schedule_request_id = db.Column(db.String(50), nullable=True)
    appointment_id = db.Column(db.Integer, db.ForeignKey('appointment.id'), nullable=True)
    doctor = db.relationship('User', backref='appointment_slots')
    appointment = db.relationship('Appointment', backref=db.backref('slot', uselist=False))
    
class Appointment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    doctor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    appointment_date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), nullable=False, default='Pending')
    patient_confirmed_finished = db.Column(db.Boolean, default=False)
    patient = db.relationship('User', foreign_keys=[patient_id])
    doctor = db.relationship('User', foreign_keys=[doctor_id])

class ResepObat(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    visit_id = db.Column(db.Integer, db.ForeignKey('visit.id'), nullable=False)
    detail_resep = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(50), nullable=False, default='Menunggu Konfirmasi Apoteker')
    visit = db.relationship('Visit', backref=db.backref('resep_obat', uselist=False, cascade="all, delete-orphan"))

# --- [PERUBAHAN] ---
class Pembayaran(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    visit_id = db.Column(db.Integer, db.ForeignKey('visit.id'), nullable=False)
    status = db.Column(db.String(50), nullable=False, default='Belum Lunas')
    tanggal_lunas = db.Column(db.DateTime, nullable=True)
    
    # Kolom biaya dipisah
    biaya_konsultasi = db.Column(db.Float, nullable=True, default=0.0)
    biaya_obat = db.Column(db.Float, nullable=True, default=0.0)
    
    visit = db.relationship('Visit', backref=db.backref('pembayaran', uselist=False, cascade="all, delete-orphan"))

    # Properti untuk menjumlahkan total secara otomatis
    @property
    def jumlah(self):
        total = 0
        if self.biaya_konsultasi:
            total += self.biaya_konsultasi
        if self.biaya_obat:
            total += self.biaya_obat
        return total

class LogAudit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    aktivitas = db.Column(db.String(255), nullable=False)
    timestamp = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    user = db.relationship('User')

class PesanSupport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subjek = db.Column(db.String(255), nullable=False)
    pesan = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(50), nullable=False, default='Baru')
    timestamp = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(timezone.utc))
    user = db.relationship('User', backref='pesan_support')

login_manager.login_view = 'login_choice'
login_manager.login_message = "Silakan login untuk mengakses halaman ini."
login_manager.login_message_category = "warning"

@login_manager.unauthorized_handler
def unauthorized():
    """Custom handler untuk unauthorized access"""
    if request.endpoint and 'staff' in request.endpoint:
        flash('Silakan login sebagai staff.', 'warning')
        return redirect(url_for('staff_login'))
    elif request.endpoint and 'patient' in request.endpoint:
        flash('Silakan login sebagai pasien.', 'warning')
        return redirect(url_for('patient_login'))
    else:
        return redirect(url_for('login_choice'))

def hash_ip(ip_address):
    """Hash IP address untuk privacy"""
    return hashlib.sha256(ip_address.encode()).hexdigest()[:16]

def check_login_attempts(username, role_type):
    """
    Cek jumlah percobaan login yang gagal
    role_type: 'staff' atau 'patient'
    """
    key = f"login_attempts_{role_type}_{username}"
    attempts = session.get(key, 0)
    
    if attempts >= 8:
        lockout_time = session.get(f"{key}_lockout_until")
        if lockout_time and datetime.now() < lockout_time:
            minutes_left = int((lockout_time - datetime.now()).total_seconds() / 60) + 1
            return False, f"Akun terkunci. Coba lagi dalam {minutes_left} menit."
        else:
            # Lockout expired, reset attempts
            session.pop(key, None)
            session.pop(f"{key}_lockout_until", None)
            return True, None
    
    return True, None

def record_failed_login(username, role_type):
    """Catat percobaan login yang gagal"""
    key = f"login_attempts_{role_type}_{username}"
    attempts = session.get(key, 0) + 1
    session[key] = attempts
    
    if attempts >= 8:
        session[f"{key}_lockout_until"] = datetime.now() + timedelta(minutes=20)
        return True  # Locked
    
    return False

def reset_login_attempts(username, role_type):
    """Reset percobaan login setelah sukses"""
    key = f"login_attempts_{role_type}_{username}"
    session.pop(key, None)
    session.pop(f"{key}_lockout_until", None)
    
    reset_rate_limiter_for_user()
    
def generate_session_token():
    """Generate token unik untuk session"""
    return secrets.token_urlsafe(32)

def validate_session_token():
    """Validasi session token untuk mencegah session hijacking"""
    stored_token = session.get('session_token')
    stored_ip = session.get('session_ip')
    current_ip = hash_ip(request.remote_addr)
    
    if not stored_token or stored_ip != current_ip:
        return False
    
    return True

def staff_required(fn):
    """Decorator untuk route yang hanya bisa diakses staff"""
    @wraps(fn)
    def decorated_view(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Silakan login terlebih dahulu.', 'warning')
            return redirect(url_for('staff_login'))
        
        if current_user.role not in ['admin', 'dokter', 'apoteker']:
            flash('Akses ditolak. Halaman ini hanya untuk staff.', 'danger')
            return redirect(url_for('patient_login'))
        
        # Validasi session token
        if not validate_session_token():
            logout_user()
            flash('Session tidak valid. Silakan login ulang.', 'warning')
            return redirect(url_for('staff_login'))
        
        return fn(*args, **kwargs)
    return decorated_view

def patient_required(fn):
    """Decorator untuk route yang hanya bisa diakses pasien"""
    @wraps(fn)
    def decorated_view(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Silakan login terlebih dahulu.', 'warning')
            return redirect(url_for('patient_login'))
        
        if current_user.role != 'pasien':
            flash('Akses ditolak. Halaman ini hanya untuk pasien.', 'danger')
            return redirect(url_for('staff_login'))
        
        # Validasi session token
        if not validate_session_token():
            logout_user()
            flash('Session tidak valid. Silakan login ulang.', 'warning')
            return redirect(url_for('patient_login'))
        
        return fn(*args, **kwargs)
    return decorated_view

def send_async_email(app, msg):
    """Fungsi untuk pengiriman email asinkron."""
    with app.app_context():
        try:
            mail.send(msg)
        except Exception as e:
            print(f"Error mengirim email: {e}")

def send_email(to, subject, template):
    """Menyiapkan dan memulai thread pengiriman email."""
    msg = Message(
        subject,
        recipients=[to],
        html=template,
        sender=app.config['MAIL_DEFAULT_SENDER']
    )
    Thread(target=send_async_email, args=(app, msg)).start()

def send_password_reset_email(user):
    """Fungsi khusus untuk mengirim email reset password."""
    token = user.get_reset_token()
    send_email(user.email,
               '[Batam Sehat] Instruksi Reset Password Anda',
               render_template('email/reset_password.html',
                               user=user, 
                               token=token, 
                               now=datetime.now()))
    
def send_email_change_confirmation(user, new_email):
    """Fungsi khusus untuk mengirim email konfirmasi ganti email."""
    token = user.get_email_change_token(new_email)
    send_email(new_email,  # Kirim ke email BARU
               '[Batam Sehat] Konfirmasi Perubahan Alamat Email Anda',
               render_template('email/confirm_email_change.html',
                               user=user, token=token, now=datetime.now()))

def send_appointment_approved_email(appointment):
    """Email saat appointment disetujui admin"""
    patient = appointment.patient
    if not patient.email:
        return
    
    subject = '[Batam Sehat] Konsultasi Anda Disetujui'
    html = render_template('email/appointment_approved.html',
                          patient=patient,
                          appointment=appointment,
                          now=datetime.now())
    send_email(patient.email, subject, html)
    

def send_appointment_reminder(appointment):
    """Email reminder H-1 sebelum konsultasi"""
    patient = appointment.patient
    if not patient.email:
        return
    
    subject = '[Batam Sehat] Reminder: Konsultasi Besok'
    html = render_template('email/appointment_reminder.html',
                          patient=patient,
                          appointment=appointment,
                          now=datetime.now())
    send_email(patient.email, subject, html)


def send_prescription_ready_email(resep):
    """Email saat resep siap diambil"""
    patient = resep.visit.patient
    if not patient.user or not patient.user.email:
        return
    
    subject = '[Batam Sehat] Resep Obat Anda Siap Diambil'
    html = render_template('email/prescription_ready.html',
                          patient=patient,
                          resep=resep,
                          now=datetime.now())
    send_email(patient.user.email, subject, html)


def send_payment_reminder(pembayaran):
    """Email reminder pembayaran"""
    patient = pembayaran.visit.patient
    if not patient.user or not patient.user.email:
        return
    
    subject = '[Batam Sehat] Tagihan Menunggu Pembayaran'
    html = render_template('email/payment_reminder.html',
                          patient=patient,
                          pembayaran=pembayaran,
                          now=datetime.now())
    send_email(patient.user.email, subject, html)


def send_account_verified_email(user):
    """Email saat akun diverifikasi admin"""
    if not user.email:
        return
    
    subject = '[Batam Sehat] Akun Anda Telah Diaktifkan'
    html = render_template('email/account_verified.html',
                          user=user,
                          now=datetime.now())
    send_email(user.email, subject, html)

def catat_log(aktivitas):
    try:
        user_id = current_user.id if current_user.is_authenticated else None
        log = LogAudit(user_id=user_id, aktivitas=aktivitas)
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error saat mencatat log: {e}")

def role_required(*roles):
    def wrapper(fn):
        @wraps(fn)
        def decorated_view(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash('Anda tidak memiliki izin untuk mengakses halaman ini.', 'danger')
                return redirect(url_for('dashboard'))
            return fn(*args, **kwargs)
        return decorated_view
    return wrapper

def verified_staff_required(fn):
    """
    Decorator untuk memastikan staf (dokter/apoteker) sudah terverifikasi
    Jika belum, redirect ke halaman pembatasan
    """
    @wraps(fn)
    def decorated_view(*args, **kwargs):
        if not current_user.is_authenticated:
            return login_manager.unauthorized()
        
        # Cek jika user adalah staf (dokter/apoteker) yang belum terverifikasi
        if current_user.role in ['dokter', 'apoteker']:
            if current_user.verification_status != 'aktif':
                flash('Fitur ini hanya dapat diakses setelah data profil Anda diverifikasi oleh Admin.', 'warning')
                return redirect(url_for('dashboard_unverified_staff'))
        
        return fn(*args, **kwargs)
    return decorated_view

def check_and_send_reminders():
    """Fungsi yang dijalankan scheduler setiap hari"""
    with app.app_context():
        tomorrow = datetime.now().date() + timedelta(days=1)
        
        # Cari appointment yang besok
        appointments = Appointment.query.filter(
            Appointment.appointment_date == tomorrow,
            Appointment.status == 'Approved'
        ).all()
        
        for appointment in appointments:
            try:
                send_appointment_reminder(appointment)
                print(f"Reminder sent to {appointment.patient.username}")
            except Exception as e:
                print(f"Error sending reminder: {e}")
                
def check_unpaid_bills():
    """Kirim reminder untuk tagihan yang belum dibayar > 3 hari"""
    with app.app_context():
        three_days_ago = datetime.now() - timedelta(days=3)
        
        unpaid_bills = Pembayaran.query.filter(
            Pembayaran.status == 'Belum Lunas'
        ).join(Visit).filter(
            Visit.visit_date <= three_days_ago
        ).all()
        
        for bill in unpaid_bills:
            try:
                send_payment_reminder(bill)
                print(f"Payment reminder sent for bill ID {bill.id}")
            except Exception as e:
                print(f"Error sending payment reminder: {e}")
                
scheduler = BackgroundScheduler()
scheduler.add_job(func=check_and_send_reminders, trigger="cron", hour=8, minute=0)  # Setiap hari jam 8 pagi
scheduler.add_job(func=check_unpaid_bills, trigger="cron", hour=16, minute=0)  # Setiap hari jam 4 sore
scheduler.start()

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Route login lama - redirect ke halaman pilihan"""
    return redirect(url_for('login_choice'))

@app.route('/login/choice')
def login_choice():
    """Halaman pemilihan tipe login"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('login_choice.html')

@app.route('/login/2fa', methods=['GET', 'POST'])
def login_2fa():
    if 'user_id_2fa' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id_2fa']
    user = User.query.get(user_id)
    if not user:
        session.pop('user_id_2fa', None)
        return redirect(url_for('login'))

    if request.method == 'POST':
        token = request.form.get('token')
        totp = pyotp.TOTP(user.otp_secret)
        if totp.verify(token):
            session.pop('user_id_2fa', None)
            login_user(user)
            catat_log(f'User {user.username} berhasil login (dengan 2FA).')
            return redirect(url_for('dashboard'))
        else:
            flash('Kode OTP tidak valid.', 'danger')
    
    return render_template('login_2fa.html') 

@app.route('/reset_password', methods=['GET', 'POST'])
@limiter.limit("3 per hour")
def reset_password_request():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        if user:
            send_password_reset_email(user)
            flash('Email untuk instruksi reset password telah dikirim. Silakan periksa inbox Anda (dan folder spam).', 'info')
            return redirect(url_for('login'))
        else:
            flash('Email tidak terdaftar di sistem kami.', 'warning')
    return render_template('forgot_password.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    user = User.verify_reset_token(token)
    if not user:
        flash('Token tidak valid atau sudah kedaluwarsa.', 'warning')
        return redirect(url_for('reset_password_request'))

    if request.method == 'POST':
        password = request.form.get('password')
        password2 = request.form.get('password2')

        if password != password2:
            flash('Password tidak cocok.', 'danger')
            return render_template('reset_password.html', token=token)

        if len(password) < 8 or not re.search(r"[a-z]", password) or not re.search(r"[A-Z]", password) or not re.search(r"\d", password):
            flash('Password harus minimal 8 karakter, mengandung huruf besar, kecil, dan angka.', 'danger')
            return render_template('reset_password.html', token=token)

        user.set_password(password)
        db.session.commit()
        catat_log(f'User {user.username} berhasil reset password via email.')
        flash('Password Anda telah berhasil diperbarui. Silakan login.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', token=token)

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("15 per hour")
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        password2 = request.form.get('password2')

        if len(password) < 8 or not re.search(r"[a-z]", password) or not re.search(r"[A-Z]", password) or not re.search(r"\d", password):
            flash('Password harus minimal 8 karakter, mengandung huruf besar, kecil, dan angka.', 'danger')
            return redirect(url_for('register'))
        
        if password != password2:
            flash('Password dan Konfirmasi Password tidak cocok.', 'danger')
            return redirect(url_for('register'))

        if User.query.filter_by(username=username).first():
            flash('Username sudah digunakan.', 'danger')
            return redirect(url_for('register'))
            
        new_user = User(
            username=username, 
            role='pasien',
            verification_status='belum diverifikasi'
        )
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        catat_log(f'Pasien baru "{username}" berhasil registrasi dan menunggu verifikasi.')
        flash('Registrasi berhasil! Silahkan Login untuk mengisi data diri anda.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    username = current_user.username
    role = current_user.role
    
    # Clear session security tokens
    session.pop('session_token', None)
    session.pop('session_ip', None)
    session.pop('login_time', None)
    
    catat_log(f'User {username} ({role}) logout.')
    logout_user()
    
    flash('Anda telah berhasil logout.', 'success')
    
    # Redirect ke halaman login yang sesuai
    if role in ['admin', 'dokter', 'apoteker']:
        return redirect(url_for('staff_login'))
    else:
        return redirect(url_for('patient_login'))


@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html')

@app.route('/request_email_change', methods=['POST'])
@login_required
def request_email_change():
    new_email = request.form.get('new_email')
    current_password = request.form.get('current_password')

    # 1. Validasi Input Dasar
    if not new_email or not current_password:
        flash('Email baru dan password saat ini wajib diisi.', 'danger')
        return redirect(url_for('profile'))

    # 2. Verifikasi Password Saat Ini
    if not current_user.check_password(current_password):
        flash('Password saat ini salah.', 'danger')
        return redirect(url_for('profile'))

    # 3. Cek apakah email sama dengan yang lama
    if new_email == current_user.email:
        flash('Email baru sama dengan email saat ini.', 'info')
        return redirect(url_for('profile'))

    # 4. Cek apakah email baru sudah digunakan oleh user lain
    existing_user = User.query.filter(User.email == new_email, User.id != current_user.id).first()
    if existing_user:
        flash('Email ini sudah digunakan oleh akun lain.', 'danger')
        return redirect(url_for('profile'))

    # 5. Kirim Email Konfirmasi (BUKAN update DB langsung)
    try:
        send_email_change_confirmation(current_user, new_email)
        catat_log(f'User {current_user.username} meminta perubahan email ke {new_email}.')
        flash('Permintaan perubahan email terkirim. Silakan periksa inbox email BARU Anda (dan folder spam) untuk link konfirmasi.', 'info')
    except Exception as e:
        flash(f'Gagal mengirim email konfirmasi: {e}', 'danger')

    return redirect(url_for('profile'))

@app.route('/confirm_email_change/<token>')
@login_required # User harus login untuk konfirmasi (keamanan tambahan)
def confirm_email_change(token):
    data = User.verify_email_change_token(token)

    if not data:
        flash('Link konfirmasi email tidak valid atau sudah kedaluwarsa.', 'warning')
        return redirect(url_for('profile'))

    user_id = data.get('user_id')
    new_email = data.get('new_email')

    # Pastikan token ini untuk user yang sedang login
    if user_id != current_user.id:
        flash('Token konfirmasi ini bukan untuk akun Anda.', 'danger')
        return redirect(url_for('profile'))

    # Safety check: pastikan email belum diambil orang lain sejak request dibuat
    existing_user = User.query.filter(User.email == new_email, User.id != current_user.id).first()
    if existing_user:
        flash(f'Email {new_email} sudah digunakan oleh akun lain. Silakan coba lagi.', 'danger')
        return redirect(url_for('profile'))

    # Update email jika semua OK
    try:
        current_user.email = new_email
        db.session.commit()
        catat_log(f'User {current_user.username} berhasil mengonfirmasi perubahan email menjadi {new_email}.')
        flash('Alamat email Anda telah berhasil diperbarui.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan saat memperbarui email: {e}', 'danger')

    return redirect(url_for('profile'))

@app.route('/apoteker/riwayat')
@login_required
@role_required('apoteker')
@verified_staff_required
def apoteker_riwayat():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('q', '')
    per_page = 10

    # Query dasar untuk mengambil data yang relevan
    query = ResepObat.query.join(Visit).filter(
        ResepObat.status != 'Menunggu Konfirmasi Apoteker'
    )
    
    riwayat_list_all = query.order_by(Visit.visit_date.desc()).all()

    # Filter manual di Python karena nama pasien dienkripsi
    if search_query:
        riwayat_list_filtered = [
            resep for resep in riwayat_list_all
            if search_query.lower() in resep.visit.patient.full_name.lower() or \
               search_query.lower() in resep.visit.doctor.username.lower()
        ]
    else:
        riwayat_list_filtered = riwayat_list_all

    # Melakukan pagination secara manual
    total = len(riwayat_list_filtered)
    total_pages = (total + per_page - 1) // per_page
    start = (page - 1) * per_page
    end = start + per_page
    riwayat_list_paginated = riwayat_list_filtered[start:end]
    
    # Menyiapkan variabel untuk dikirim ke template
    has_prev = page > 1
    prev_num = page - 1
    has_next = page < total_pages
    next_num = page + 1

    return render_template('apoteker_riwayat.html', 
                           riwayat_list=riwayat_list_paginated, 
                           search_query=search_query,
                           page=page,
                           total_pages=total_pages,
                           has_prev=has_prev,
                           prev_num=prev_num,
                           has_next=has_next,
                           next_num=next_num)
    
    from flask_sqlalchemy.pagination import Pagination
    pagination = Pagination(query=None, page=page, per_page=per_page, total=total, items=riwayat_list_paginated)


    return render_template('apoteker_riwayat.html', 
                           riwayat_list=riwayat_list_paginated, 
                           pagination=pagination,
                           search_query=search_query)

@app.route('/complete_profile', methods=['GET', 'POST'])
@login_required
def complete_profile():
    if current_user.is_profile_complete and current_user.verification_status != 'belum diverifikasi':
         return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nik = request.form.get('nik')
        full_name = request.form.get('full_name')
        address = request.form.get('address')
        contact = request.form.get('contact')
        date_of_birth = request.form.get('date_of_birth')

        if not re.match(r'^\d+$', nik):
            flash('NIK tidak valid. Harap masukkan NIK yang hanya terdiri dari angka.', 'danger')
            return render_template('complete_profile.html', current_email=current_user.email, user_role=current_user.role)

        if not re.match(r'^[a-zA-Z\s]+$', full_name):
            flash('Nama Lengkap tidak valid. Harap masukkan nama yang hanya terdiri dari huruf dan spasi.', 'danger')
            return render_template('complete_profile.html', current_email=current_user.email, user_role=current_user.role)

        if current_user.role == 'pasien':
            email = request.form.get('email')
            if not email:
                flash('Email wajib diisi.', 'danger')
                return render_template('complete_profile.html', current_email=current_user.email, user_role=current_user.role)

            existing_user_email = User.query.filter(User.email == email, User.id != current_user.id).first()
            if existing_user_email:
                flash('Email ini sudah digunakan oleh akun lain.', 'danger')
                return render_template('complete_profile.html', current_email=current_user.email, user_role=current_user.role)
            current_user.email = email

        try:
            current_user.nik_encrypted = fernet.encrypt(nik.encode())
            current_user.full_name_encrypted = fernet.encrypt(full_name.encode())
            current_user.contact_encrypted = fernet.encrypt(contact.encode())
            current_user.date_of_birth_encrypted = fernet.encrypt(date_of_birth.encode())
            current_user.address_encrypted = fernet.encrypt(address.encode())

            if current_user.role == 'pasien':
                patient_record = current_user.patient_record
                if patient_record:
                    patient_record.full_name_encrypted = current_user.full_name_encrypted
                    patient_record.date_of_birth_encrypted = current_user.date_of_birth_encrypted
                    patient_record.address_encrypted = current_user.address_encrypted

            if current_user.role in ['dokter', 'apoteker']:
                current_user.verification_status = 'menunggu verifikasi data'
            if current_user.role == 'admin':
                 if not current_user.is_active_db:
                     current_user.is_active_db = True
                 current_user.verification_status = 'aktif'

            db.session.commit()
            catat_log(f'User {current_user.username} melengkapi data profil.')
            flash('Profil Anda berhasil diperbarui!', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f"Terjadi kesalahan saat menyimpan profil: {e}", "danger")
            return render_template('complete_profile.html', current_email=current_user.email, user_role=current_user.role)

    return render_template('complete_profile.html', current_email=current_user.email, user_role=current_user.role)

@app.route('/patient/<int:patient_id>/upload_photos', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'pasien', 'dokter', 'apoteker')
def upload_patient_photos(patient_id):
    """Upload foto profil dan KTP pasien"""
    patient = Patient.query.get_or_404(patient_id)
    
    if current_user.role == 'pasien':
        if patient.user_id != current_user.id:
            flash('Anda tidak memiliki akses untuk mengupload foto pasien lain.', 'danger')
            return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        uploaded_files = []
        
        if 'profile_photo' in request.files:
            file = request.files['profile_photo']
            if file and file.filename and allowed_file(file.filename):
                if patient.profile_photo:
                    old_path = os.path.join(app.config['UPLOAD_FOLDER'], patient.profile_photo)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                
                filename = secure_filename(f"profile_{patient.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                
                # ========== PERBAIKAN: Gunakan forward slash ==========
                filepath = f"profiles/{filename}"  # Gunakan f-string dengan /
                # atau alternatif: filepath = str(Path('profiles') / filename)
                
                full_path = os.path.join(app.config['UPLOAD_FOLDER'], filepath)
                file.save(full_path)
                patient.profile_photo = filepath
                uploaded_files.append('Foto Profil')
        
        if 'ktp_photo' in request.files:
            file = request.files['ktp_photo']
            if file and file.filename and allowed_file(file.filename):
                if patient.ktp_photo:
                    old_path = os.path.join(app.config['UPLOAD_FOLDER'], patient.ktp_photo)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                
                filename = secure_filename(f"ktp_{patient.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                
                # ========== PERBAIKAN: Gunakan forward slash ==========
                filepath = f"ktp/{filename}"  # Gunakan f-string dengan /
                
                full_path = os.path.join(app.config['UPLOAD_FOLDER'], filepath)
                file.save(full_path)
                patient.ktp_photo = filepath
                uploaded_files.append('Foto KTP')
        
        if uploaded_files:
            db.session.commit()
            catat_log(f'User {current_user.username} mengupload {", ".join(uploaded_files)} untuk pasien {patient.full_name}.')
            flash(f'{", ".join(uploaded_files)} berhasil diupload!', 'success')
        else:
            flash('Tidak ada file yang diupload.', 'warning')
        
        if current_user.role == 'pasien':
            return redirect(url_for('profile'))
        else:
            return redirect(url_for('patient_detail', patient_id=patient.id))
    
    return render_template('upload_patient_photos.html', patient=patient)

@app.route('/patient/<int:patient_id>/upload_document', methods=['POST'])
@login_required
@role_required('admin', 'dokter')
def upload_medical_document(patient_id):
    """Upload dokumen medis (hasil lab, X-Ray, dll)"""
    patient = Patient.query.get_or_404(patient_id)
    
    if 'document_file' not in request.files:
        flash('File tidak ditemukan.', 'danger')
        return redirect(url_for('patient_detail', patient_id=patient.id))
    
    file = request.files['document_file']
    document_type = request.form.get('document_type')
    description = request.form.get('description', '')
    
    if not file or not file.filename:
        flash('Silakan pilih file yang akan diupload.', 'warning')
        return redirect(url_for('patient_detail', patient_id=patient.id))
    
    if not allowed_file(file.filename):
        flash('Format file tidak didukung. Gunakan PNG, JPG, JPEG, GIF, atau PDF.', 'danger')
        return redirect(url_for('patient_detail', patient_id=patient.id))
    
    filename = secure_filename(f"doc_{patient.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
    
    # ========== PERBAIKAN: Gunakan forward slash ==========
    filepath = f"documents/{filename}"  # Gunakan f-string dengan /
    
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], filepath)
    file.save(full_path)
    
    file_size = os.path.getsize(full_path)
    
    new_document = PatientDocument(
        patient_id=patient.id,
        document_type=document_type,
        file_name=file.filename,
        file_path=filepath,
        file_size=file_size,
        uploaded_by=current_user.id,
        description=description
    )
    db.session.add(new_document)
    db.session.commit()
    
    catat_log(f'{current_user.role.capitalize()} {current_user.username} mengupload dokumen {document_type} untuk pasien {patient.full_name}.')
    flash(f'Dokumen "{file.filename}" berhasil diupload!', 'success')
    return redirect(url_for('patient_detail', patient_id=patient.id))

@app.route('/document/<int:document_id>/delete', methods=['POST'])
@login_required
@role_required('admin', 'dokter')
def delete_medical_document(document_id):
    """Hapus dokumen medis"""
    document = PatientDocument.query.get_or_404(document_id)
    patient_id = document.patient_id
    
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], document.file_path)
    if os.path.exists(full_path):
        os.remove(full_path)
    
    db.session.delete(document)
    db.session.commit()
    
    catat_log(f'{current_user.role.capitalize()} {current_user.username} menghapus dokumen "{document.file_name}".')
    flash('Dokumen berhasil dihapus.', 'success')
    return redirect(url_for('patient_detail', patient_id=patient_id))

@app.route('/setup_2fa', methods=['GET', 'POST'])
@login_required
def setup_2fa():
    if request.method == 'POST':
        token = request.form.get('token')
        secret = session.get('otp_secret_setup')
        if not secret:
            return redirect(url_for('profile'))
            
        totp = pyotp.TOTP(secret)
        if totp.verify(token):
            current_user.otp_secret = secret
            db.session.commit()
            session.pop('otp_secret_setup', None)
            catat_log(f'User {current_user.username} berhasil mengaktifkan 2FA.')
            flash('Autentikasi Dua Faktor berhasil diaktifkan!', 'success')
            return redirect(url_for('profile'))
        else:
            flash('Kode verifikasi tidak valid. Silakan coba lagi.', 'danger')
            session.pop('otp_secret_setup', None)
            return redirect(url_for('setup_2fa'))

    if 'otp_secret_setup' not in session:
        session['otp_secret_setup'] = pyotp.random_base32()
    
    secret = session['otp_secret_setup']
    provisioning_uri = pyotp.totp.TOTP(secret).provisioning_uri(
        name=current_user.username,
        issuer_name='Batam Sehat App'
    )
    
    qr_img = qrcode.make(provisioning_uri)
    buffered = io.BytesIO()
    qr_img.save(buffered, format="PNG")
    qr_code_b64 = base64.b64encode(buffered.getvalue()).decode()

    return render_template('setup_2fa.html', qr_code=qr_code_b64, secret=secret)

@app.route('/staff/upload_photos', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'dokter', 'apoteker')
def upload_staff_photos():
    """Upload foto profil dan KTP untuk staff (admin, dokter, apoteker)"""
    
    if request.method == 'POST':
        uploaded_files = []
        
        if 'profile_photo' in request.files:
            file = request.files['profile_photo']
            if file and file.filename:
                is_valid, error_msg = validate_image_file(file, max_size_mb=5)
                if not is_valid:
                    flash(f'Foto Profil: {error_msg}', 'danger')
                else:
                    if current_user.profile_photo:
                        old_path = os.path.join(app.config['UPLOAD_FOLDER'], current_user.profile_photo)
                        if os.path.exists(old_path):
                            try:
                                os.remove(old_path)
                            except Exception as e:
                                print(f"Error menghapus foto lama: {e}")
                    
                    filename = secure_filename(f"staff_profile_{current_user.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                    
                    # ========== PERBAIKAN: Gunakan forward slash ==========
                    filepath = f"profiles/{filename}"  # Gunakan f-string dengan /
                    
                    full_path = os.path.join(app.config['UPLOAD_FOLDER'], filepath)
                    
                    try:
                        file.save(full_path)
                        current_user.profile_photo = filepath
                        uploaded_files.append('Foto Profil')
                    except Exception as e:
                        flash(f'Gagal menyimpan foto profil: {e}', 'danger')
        
        if 'ktp_photo' in request.files:
            file = request.files['ktp_photo']
            if file and file.filename:
                is_valid, error_msg = validate_image_file(file, max_size_mb=5)
                if not is_valid:
                    flash(f'Foto KTP: {error_msg}', 'danger')
                else:
                    if current_user.ktp_photo:
                        old_path = os.path.join(app.config['UPLOAD_FOLDER'], current_user.ktp_photo)
                        if os.path.exists(old_path):
                            try:
                                os.remove(old_path)
                            except Exception as e:
                                print(f"Error menghapus foto lama: {e}")
                    
                    filename = secure_filename(f"staff_ktp_{current_user.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                    
                    # ========== PERBAIKAN: Gunakan forward slash ==========
                    filepath = f"ktp/{filename}"  # Gunakan f-string dengan /
                    
                    full_path = os.path.join(app.config['UPLOAD_FOLDER'], filepath)
                    
                    try:
                        file.save(full_path)
                        current_user.ktp_photo = filepath
                        uploaded_files.append('Foto KTP')
                    except Exception as e:
                        flash(f'Gagal menyimpan foto KTP: {e}', 'danger')
        
        if uploaded_files:
            try:
                db.session.commit()
                catat_log(f'{current_user.role.capitalize()} {current_user.username} mengupload {", ".join(uploaded_files)}.')
                flash(f'{", ".join(uploaded_files)} berhasil diupload!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Gagal menyimpan ke database: {e}', 'danger')
        else:
            flash('Tidak ada file yang diupload.', 'warning')
        
        return redirect(url_for('profile'))
    
    return render_template('upload_staff_photos.html')

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    if not current_user.is_active_db:
        flash('Akun Anda saat ini tidak aktif. Akses terbatas hanya untuk menghubungi Admin.', 'warning')
        return render_template('dashboard_inactive.html')
    
    if not current_user.is_profile_complete:
        if current_user.verification_status not in ['belum diverifikasi', 'menunggu verifikasi data']:
             current_user.verification_status = 'belum diverifikasi'
             db.session.commit()

        flash('Harap lengkapi data diri Anda sebelum melanjutkan.', 'info')
        return redirect(url_for('complete_profile'))

    if current_user.role in ['dokter', 'apoteker'] and current_user.verification_status == 'menunggu verifikasi data':
        flash('Data profil Anda sedang ditinjau oleh Admin. Akses fitur terbatas.', 'warning')
        return redirect(url_for('dashboard_unverified_staff'))

    if current_user.role == 'pasien' and current_user.verification_status == 'belum diverifikasi':
        flash('Akun Anda sedang menunggu verifikasi oleh Admin. Harap tunggu.', 'warning')

    if current_user.role == 'admin':
        appointments = Appointment.query.filter(Appointment.status != 'Finished').order_by(Appointment.appointment_date.asc()).all()
        return render_template('dashboard_admin.html', appointments=appointments)

    elif current_user.role == 'dokter':
        appointments = Appointment.query.filter_by(doctor_id=current_user.id, status='Approved').order_by(Appointment.appointment_date.asc()).all()
        return render_template('dashboard.html', appointments=appointments)

    elif current_user.role == 'pasien':
        if current_user.verification_status != 'aktif':
             flash('Akun Anda sedang menunggu verifikasi oleh Admin.', 'warning')
        appointments = Appointment.query.filter_by(patient_id=current_user.id).order_by(Appointment.appointment_date.desc()).all()
        return render_template('dashboard_pasien.html', appointments=appointments)

    elif current_user.role == 'apoteker':
        return redirect(url_for('apoteker_dashboard'))

    return render_template('dashboard.html')

@app.route('/manage_patients')
@login_required
@role_required('admin', 'dokter', 'apoteker')
@verified_staff_required
def manage_patients():
    search_query = request.args.get('q', '')

    if current_user.role == 'admin':
        query_base = Patient.query
    else: 
        patient_ids = db.session.query(Appointment.patient_id).filter_by(doctor_id=current_user.id).distinct()
        query_base = Patient.query.filter(Patient.user_id.in_(patient_ids))

    if search_query:
        query_base = query_base.filter(Patient.medical_record_number.like(f'%{search_query}%'))

    all_patients = query_base.all()
    
    patients = sorted(all_patients, key=lambda patient: patient.full_name.lower())
    
    return render_template('manage_patients.html', patients=patients, search_query=search_query)


@app.route('/patient/add', methods=['POST'])
@login_required
@role_required('admin')
def add_patient():
    full_name_encrypted = fernet.encrypt(request.form.get('full_name').encode())
    date_of_birth_encrypted = fernet.encrypt(request.form.get('date_of_birth').encode())
    address_encrypted = fernet.encrypt(request.form.get('address').encode())
    new_patient = Patient(
        medical_record_number=request.form.get('medical_record_number'),
        full_name_encrypted=full_name_encrypted,
        date_of_birth_encrypted=date_of_birth_encrypted,
        address_encrypted=address_encrypted
    )
    db.session.add(new_patient)
    db.session.commit()
    catat_log(f'Admin {current_user.username} menambahkan pasien baru (manual): {request.form.get("full_name")}.')
    flash('Pasien baru berhasil ditambahkan!', 'success')
    return redirect(url_for('manage_patients'))

@app.route('/patient/<int:patient_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_patient(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    if request.method == 'POST':
        patient.medical_record_number = request.form.get('medical_record_number')
        patient.full_name_encrypted = fernet.encrypt(request.form.get('full_name').encode())
        patient.date_of_birth_encrypted = fernet.encrypt(request.form.get('date_of_birth').encode())
        patient.address_encrypted = fernet.encrypt(request.form.get('address').encode())
        db.session.commit()
        catat_log(f'Admin {current_user.username} mengubah data pasien: {patient.full_name}.')
        flash('Data pasien berhasil diperbarui!', 'success')
        return redirect(url_for('manage_patients'))
    return render_template('edit_patient.html', patient=patient)
    
@app.route('/patient/<int:patient_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def delete_patient(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    patient_name = patient.full_name
    db.session.delete(patient)
    db.session.commit()
    catat_log(f'Admin {current_user.username} menghapus pasien: {patient_name}.')
    flash('Data pasien berhasil dihapus.', 'success')
    return redirect(url_for('manage_patients'))

@app.route('/patient/<int:patient_id>')
@login_required
@role_required('admin', 'dokter')
@verified_staff_required
def patient_detail(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    appointment_id = request.args.get('appointment_id', type=int)
    return render_template('patient_detail.html', patient=patient, appointment_id=appointment_id)

# --- [PERUBAHAN] ---
@app.route('/patient/<int:patient_id>/add_visit', methods=['POST'])
@login_required
@role_required('dokter')
@verified_staff_required
def add_visit(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    appointment_id = request.form.get('appointment_id', type=int)
    biaya_konsultasi_str = request.form.get('biaya_konsultasi', '0').replace('.', '').replace(',', '')

    try:
        biaya_konsultasi = float(biaya_konsultasi_str)
    except ValueError:
        flash('Biaya konsultasi tidak valid.', 'danger')
        return redirect(url_for('patient_detail', patient_id=patient.id, appointment_id=appointment_id))

    new_visit = Visit(
        patient_id=patient.id, 
        doctor_id=current_user.id,
        appointment_id=appointment_id
    )
    db.session.add(new_visit)
    db.session.flush()
    
    new_soap_note = SoapNote(
        subjective=request.form.get('subjective'),
        objective=request.form.get('objective'),
        assessment=request.form.get('assessment'),
        plan=request.form.get('plan'),
        visit_id=new_visit.id
    )
    db.session.add(new_soap_note)

    # Buat entri pembayaran dengan biaya konsultasi dari dokter
    pembayaran = Pembayaran(visit_id=new_visit.id, biaya_konsultasi=biaya_konsultasi)
    db.session.add(pembayaran)

    detail_resep = request.form.get('detail_resep')
    if detail_resep:
        new_resep = ResepObat(
            visit_id=new_visit.id, 
            detail_resep=detail_resep, 
            status='Menunggu Konfirmasi Apoteker'
        )
        db.session.add(new_resep)

    if appointment_id:
        appointment = Appointment.query.get(appointment_id)
        if appointment:
            appointment.status = 'Finished'
            catat_log(f'Dokter menyelesaikan konsultasi ID {appointment_id}.')

    catat_log(f'Dokter {current_user.username} menambahkan rekam medis dan biaya konsultasi untuk pasien {patient.full_name}.')
    db.session.commit()
    flash('Rekam medis berhasil disimpan dan konsultasi ditandai selesai.', 'success')
    return redirect(url_for('dashboard'))

# --- Rute Jadwal ---
@app.route('/schedule', methods=['GET', 'POST'])
@login_required
@role_required('dokter')
@verified_staff_required
def doctor_schedule():
    if request.method == 'POST':
        schedule_date_str = request.form.get('schedule_date')
        first_slot_start_str = request.form.get('first_slot_start')
        consultation_duration = int(request.form.get('consultation_duration', 60))
        
        # Ambil break duration (bisa dari dropdown atau custom)
        break_duration_select = request.form.get('break_duration')
        if break_duration_select == 'custom':
            break_duration = int(request.form.get('custom_break_minutes', 0))
        else:
            break_duration = int(break_duration_select)
            
        max_slots = int(request.form.get('max_slots', 5))

        # [FITUR BARU] Ambil data waktu istirahat tetap
        enable_break1 = request.form.get('enable_break1') == 'on'
        enable_break2 = request.form.get('enable_break2') == 'on'
        
        break_times = []
        if enable_break1:
            break1_start = request.form.get('break1_start')
            break1_end = request.form.get('break1_end')
            if break1_start and break1_end:
                break_times.append((break1_start, break1_end))
        
        if enable_break2:
            break2_start = request.form.get('break2_start')
            break2_end = request.form.get('break2_end')
            if break2_start and break2_end:
                break_times.append((break2_start, break2_end))

        if not schedule_date_str or not first_slot_start_str:
            flash('Tanggal dan jam mulai slot pertama harus diisi.', 'danger')
            return redirect(url_for('doctor_schedule'))

        try:
            schedule_date = datetime.strptime(schedule_date_str, '%Y-%m-%d').date()
            first_slot_start = datetime.strptime(first_slot_start_str, '%H:%M').time()

            # [FITUR BARU] Konversi break times ke format menit
            break_ranges = []
            for break_start_str, break_end_str in break_times:
                break_start_time = datetime.strptime(break_start_str, '%H:%M').time()
                break_end_time = datetime.strptime(break_end_str, '%H:%M').time()
                break_start_min = break_start_time.hour * 60 + break_start_time.minute
                break_end_min = break_end_time.hour * 60 + break_end_time.minute
                break_ranges.append((break_start_min, break_end_min))

            # Generate slot dengan jeda DAN skip break time
            request_id = str(uuid.uuid4())
            new_slots_batch = []
            
            # Konversi jam mulai ke menit
            current_minutes = first_slot_start.hour * 60 + first_slot_start.minute

            slots_created = 0
            safety_counter = 0  # Cegah infinite loop
            max_attempts = 200  # Batas maksimal iterasi

            while slots_created < max_slots and safety_counter < max_attempts:
                safety_counter += 1
                
                # Hitung waktu mulai dan selesai slot ini
                slot_start_minutes = current_minutes
                slot_end_minutes = current_minutes + consultation_duration

                # [FITUR BARU] Cek apakah slot TIDAK BENTROK dengan break time
                is_conflict = False
                for break_start, break_end in break_ranges:
                    # Slot bentrok jika:
                    # 1. Slot mulai di tengah break time
                    # 2. Slot selesai di tengah break time
                    # 3. Slot "menutupi" break time (mulai sebelum, selesai sesudah)
                    if (break_start <= slot_start_minutes < break_end) or \
                       (break_start < slot_end_minutes <= break_end) or \
                       (slot_start_minutes <= break_start and slot_end_minutes >= break_end):
                        is_conflict = True
                        # Skip ke akhir break time ini
                        current_minutes = break_end + break_duration
                        break
                
                if is_conflict:
                    continue  # Coba slot berikutnya

                # Konversi kembali ke datetime
                slot_start_time = datetime.combine(
                    schedule_date,
                    (datetime.min + timedelta(minutes=slot_start_minutes)).time()
                )
                slot_end_time = datetime.combine(
                    schedule_date,
                    (datetime.min + timedelta(minutes=slot_end_minutes)).time()
                )

                new_slot = AppointmentSlot(
                    doctor_id=current_user.id,
                    start_time=slot_start_time,
                    end_time=slot_end_time,
                    status='Available',
                    verification_status='Pending',
                    schedule_request_id=request_id
                )
                new_slots_batch.append(new_slot)
                slots_created += 1
                
                # Update waktu untuk slot berikutnya: tambah durasi konsultasi + jeda
                current_minutes = slot_end_minutes + break_duration

            if slots_created < max_slots:
                flash(f'Perhatian: Hanya {slots_created} slot yang berhasil dibuat (dari {max_slots} yang diminta) karena konflik dengan waktu istirahat.', 'warning')

            db.session.add_all(new_slots_batch)
            db.session.commit()

            catat_log(f'Dokter {current_user.username} mengajukan {len(new_slots_batch)} slot jadwal baru untuk tanggal {schedule_date_str}.')
            flash(f'{len(new_slots_batch)} slot jadwal baru berhasil diajukan dan menunggu persetujuan Admin.', 'info')

        except ValueError as e:
            db.session.rollback()
            flash(f'Error validasi: {e}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi error sistem saat mengajukan jadwal: {e}', 'danger')
        
        return redirect(url_for('doctor_schedule'))

    # ========== BAGIAN GET REQUEST (INI YANG KURANG) ==========
    all_slots = AppointmentSlot.query.filter(
        AppointmentSlot.doctor_id == current_user.id
    ).order_by(AppointmentSlot.start_time.asc()).all()

    schedule_groups = {}
    for slot in all_slots:
        req_id = slot.schedule_request_id
        if req_id not in schedule_groups:
            schedule_groups[req_id] = {
                'slots': [],
                'verification_status': slot.verification_status,
                'date': slot.start_time.strftime('%d-%m-%Y')
            }
        schedule_groups[req_id]['slots'].append(slot)
    
    now = datetime.utcnow()
    for req_id, group_data in schedule_groups.items():
        if group_data['verification_status'] == 'Approved':
            all_slots_in_group = group_data['slots']
            all_slots_finished = True
            
            for slot in all_slots_in_group:
                if slot.status == 'Available' and slot.start_time > now:
                    all_slots_finished = False
                    break
            
            if all_slots_finished:
                group_data['verification_status'] = 'Selesai'
    
    return render_template('doctor_schedule.html', schedule_groups=schedule_groups)

@app.route('/schedule/slot/<int:slot_id>/delete', methods=['POST'])
@login_required
@role_required('dokter')
def delete_schedule_slot(slot_id):
    slot = AppointmentSlot.query.get_or_404(slot_id)
    if slot.doctor_id != current_user.id or slot.status == 'Booked':
        flash('Anda tidak bisa menghapus slot yang sudah dipesan atau bukan milik Anda.', 'danger')
        return redirect(url_for('doctor_schedule'))

    db.session.delete(slot)
    db.session.commit()
    catat_log(f'Dokter {current_user.username} menghapus slot jadwal ID {slot.id}.')
    flash('Slot jadwal berhasil dihapus.', 'success')
    return redirect(url_for('doctor_schedule'))


@app.route('/request_appointment', methods=['GET', 'POST'])
@login_required
@role_required('pasien')
def request_appointment():
    if current_user.verification_status != 'aktif' or not current_user.is_profile_complete:
        flash('Akun Anda belum diverifikasi atau profil Anda belum lengkap.', 'warning')
        return redirect(url_for('dashboard'))
    
    if not current_user.patient_record:
        flash('Data rekam medis Anda tidak ditemukan. Harap hubungi admin.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        # ... (Bagian POST tidak ada perubahan, biarkan seperti semula)
        slot_id = request.form.get('slot_id')
        reason = request.form.get('reason')
        slot = AppointmentSlot.query.get(slot_id)

        if not slot or slot.status != 'Available' or slot.start_time <= datetime.utcnow():
            flash('Slot yang Anda pilih tidak valid, sudah dipesan, atau sudah lewat waktu.', 'danger')
            return redirect(url_for('request_appointment'))

        new_appointment = Appointment(
            patient_id=current_user.id,
            doctor_id=slot.doctor_id,
            appointment_date=slot.start_time.date(),
            reason=reason,
            status='Approved'
        )
        db.session.add(new_appointment)
        
        slot.status = 'Booked'
        db.session.flush()
        slot.appointment_id = new_appointment.id
        
        db.session.commit()
        
        catat_log(f'Pasien {current_user.username} membuat janji temu pada slot ID {slot.id}.')
        flash('Janji temu Anda berhasil dibuat.', 'success')
        return redirect(url_for('my_appointments'))
    
    # --- (Bagian GET di bawah ini yang kita ubah) ---
    doctors = User.query.filter_by(role='dokter').all()
    selected_doctor_id = request.args.get('doctor_id', type=int)
    
    available_slots = []
    if selected_doctor_id:
        # [PERBAIKAN] Tambahkan filter 'start_time > datetime.utcnow()'
        available_slots = AppointmentSlot.query.filter(
            AppointmentSlot.doctor_id == selected_doctor_id,
            AppointmentSlot.status == 'Available', 
            AppointmentSlot.verification_status == 'Approved',
            AppointmentSlot.start_time > datetime.utcnow()  # <-- BARIS INI DITAMBAHKAN
        ).order_by(AppointmentSlot.start_time.asc()).all()
    
    return render_template('request_appointment.html', doctors=doctors, slots=available_slots, selected_doctor_id=selected_doctor_id)
    
@app.route('/manage_schedules')
@login_required
@role_required('admin')
def manage_schedules():
    # 1. Ambil pengajuan jadwal BARU
    pending_slots = AppointmentSlot.query.filter_by(verification_status='Pending').order_by(AppointmentSlot.start_time.asc()).all()
    schedule_requests = {}
    for slot in pending_slots:
        key = (slot.doctor_id, slot.schedule_request_id)
        if key not in schedule_requests:
            schedule_requests[key] = {
                'doctor': slot.doctor.username,
                'date': slot.start_time.strftime('%d-%m-%Y'),
                'slots': [],
                'request_id': slot.schedule_request_id
            }
        schedule_requests[key]['slots'].append(slot)
    
    # 2. [BARU] Ambil pengajuan SELESAI
    pending_finish_slots = AppointmentSlot.query.filter(
        AppointmentSlot.verification_status == 'Menunggu Penyelesaian'
    ).order_by(AppointmentSlot.start_time.asc()).all()
    
    finish_requests = {}
    for slot in pending_finish_slots:
        key = (slot.doctor_id, slot.schedule_request_id)
        if key not in finish_requests:
            # Ambil semua slot di grup ini
            all_slots_in_group = AppointmentSlot.query.filter_by(
                schedule_request_id=slot.schedule_request_id
            ).all()
            
            # Hitung sisa slot yang available
            sisa_available = sum(1 for s in all_slots_in_group if s.status == 'Available' and s.verification_status == 'Menunggu Penyelesaian')
            
            finish_requests[key] = {
                'doctor': slot.doctor.username,
                'date': slot.start_time.strftime('%d-%m-%Y'),
                'total_slots': len(all_slots_in_group),
                'sisa_available': sisa_available,
                'request_id': slot.schedule_request_id
            }
    
    return render_template(
        'manage_schedule.html', 
        schedule_requests=schedule_requests.values(),
        finish_requests=finish_requests.values() # Kirim data baru
    )

@app.route('/schedules/<string:request_id>/update', methods=['POST'])
@login_required
@role_required('admin')
def update_schedule_status(request_id):
    new_status = request.form.get('status')
    slots_to_update = AppointmentSlot.query.filter_by(schedule_request_id=request_id).all()
    
    if not slots_to_update:
        flash('Pengajuan jadwal tidak ditemukan.', 'danger')
        return redirect(url_for('manage_schedules'))

    for slot in slots_to_update:
        if new_status == 'Rejected':
            db.session.delete(slot)
        else:
            slot.verification_status = new_status
            
    db.session.commit()
    catat_log(f'Admin {current_user.username} mengubah status pengajuan jadwal ID {request_id} menjadi {new_status}.')
    flash(f'Pengajuan jadwal telah berhasil diubah statusnya menjadi {new_status}.', 'success')
    return redirect(url_for('manage_schedules'))

@app.route('/schedules/<string:request_id>/confirm_finish', methods=['POST'])
@login_required
@role_required('admin')
def confirm_finish_schedule(request_id):
    # Ambil semua slot yang menunggu konfirmasi
    slots_to_confirm_finish = AppointmentSlot.query.filter_by(
        schedule_request_id=request_id,
        verification_status='Menunggu Penyelesaian'
    ).all()
    
    if not slots_to_confirm_finish:
        flash('Pengajuan penyelesaian jadwal tidak ditemukan.', 'danger')
        return redirect(url_for('manage_schedules'))

    try:
        for slot in slots_to_confirm_finish:
            # Selesaikan semua slot di grup ini
            slot.verification_status = 'Selesai'
                
        db.session.commit()
        catat_log(f'Admin {current_user.username} mengonfirmasi penyelesaian jadwal ID {request_id}.')
        flash(f'Pengajuan jadwal telah dikonfirmasi Selesai.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi error saat konfirmasi penyelesaian: {e}', 'danger')

    return redirect(url_for('manage_schedules'))

@app.route('/my_appointments')
@login_required
@role_required('pasien')
def my_appointments():
    appointments = Appointment.query.filter_by(patient_id=current_user.id).order_by(Appointment.appointment_date.desc()).all()
    return render_template('my_appointments.html', appointments=appointments)

@app.route('/view_appointments')
@login_required
@role_required('dokter')
@verified_staff_required
def view_appointments():
    appointments = Appointment.query.filter_by(doctor_id=current_user.id, status='Approved').order_by(Appointment.appointment_date.desc()).all()
    return render_template('view_appointments.html', appointments=appointments)

@app.route('/manage_appointments')
@login_required
@role_required('admin')
def manage_appointments():
    appointments = Appointment.query.order_by(Appointment.appointment_date.asc()).all()
    return render_template('manage_appointments.html', appointments=appointments)

@app.route('/appointment/<int:appointment_id>/update', methods=['POST'])
@login_required
@role_required('admin')
def update_appointment_status(appointment_id):
    appointment = Appointment.query.get_or_404(appointment_id)
    new_status = request.form.get('status')
    appointment.status = new_status
    db.session.commit()
    
    catat_log(f'Admin {current_user.username} mengubah status appointment ID {appointment.id} menjadi {new_status}.')
    
    # ✨ TAMBAHAN BARU: Kirim email jika disetujui
    if new_status == 'Approved':
        try:
            send_appointment_approved_email(appointment)
        except Exception as e:
            print(f"Error sending email: {e}")
    
    flash(f'Status permintaan telah diubah menjadi {new_status}.', 'success')
    return redirect(url_for('manage_appointments'))

@app.route('/appointment/<int:appointment_id>/confirm_finished', methods=['POST'])
@login_required
@role_required('pasien')
def confirm_finished(appointment_id):
    appointment = Appointment.query.get_or_404(appointment_id)
    if appointment.patient_id != current_user.id:
        return redirect(url_for('my_appointments'))
    
    appointment.patient_confirmed_finished = True
    db.session.commit()
    flash('Konsultasi telah Anda konfirmasi selesai.', 'success')
    return redirect(url_for('my_appointments'))

# (Salin kode ini dan tempelkan di app.py sekitar baris 760)

@app.route('/appointment/<int:appointment_id>/confirm_no_meds', methods=['POST'])
@login_required
@role_required('pasien')
def confirm_finished_no_meds(appointment_id):
    appointment = Appointment.query.get_or_404(appointment_id)
    if appointment.patient_id != current_user.id or appointment.status != 'Finished':
        flash('Aksi tidak valid.', 'danger')
        return redirect(url_for('my_appointments'))

    # Tandai janji temu selesai
    appointment.patient_confirmed_finished = True
    
    visit = appointment.visit
    if visit:
        # 1. Batalkan resep jika ada
        resep = visit.resep_obat
        if resep:
            resep.status = 'Dibatalkan Pasien' # Status baru untuk riwayat
        
        # 2. Nol-kan biaya obat di tagihan
        pembayaran = visit.pembayaran
        if pembayaran:
            pembayaran.biaya_obat = 0.0
    
    db.session.commit()
    catat_log(f'Pasien {current_user.username} konfirmasi selesai (tanpa obat) untuk appt ID {appointment.id}.')
    flash('Konsultasi telah dikonfirmasi selesai (tanpa pengambilan obat). Biaya obat telah dibatalkan.', 'success')
    return redirect(url_for('my_appointments'))

@app.route('/apoteker/dashboard')
@login_required
@role_required('apoteker')
@verified_staff_required
def apoteker_dashboard():
    # Mengambil resep yang masih menunggu konfirmasi harga dari apoteker
    resep_list = db.session.query(ResepObat).filter(
        ResepObat.status == 'Menunggu Konfirmasi Apoteker'
    ).join(Visit).order_by(Visit.visit_date.desc()).all()
    return render_template('dashboard_apoteker.html', resep_list=resep_list)

# --- [FUNGSI BARU] ---
@app.route('/resep/<int:resep_id>/konfirmasi', methods=['POST'])
@login_required
@role_required('apoteker')
@verified_staff_required
def konfirmasi_resep_apoteker(resep_id):
    resep = ResepObat.query.get_or_404(resep_id)
    biaya_obat_str = request.form.get(f'biaya_obat_{resep_id}', '0').replace('.', '').replace(',', '')

    try:
        biaya_obat = float(biaya_obat_str)
    except ValueError:
        flash('Biaya obat tidak valid.', 'danger')
        return redirect(url_for('apoteker_dashboard'))

    pembayaran = Pembayaran.query.filter_by(visit_id=resep.visit_id).first()
    if not pembayaran:
        pembayaran = Pembayaran(visit_id=resep.visit_id)
        db.session.add(pembayaran)

    pembayaran.biaya_obat = biaya_obat
    resep.status = 'Menunggu Pembayaran'
    
    catat_log(f'Apoteker {current_user.username} mengonfirmasi biaya obat untuk resep ID: {resep.id}.')
    db.session.commit()
    
    # ✨ TAMBAHAN BARU: Kirim email ke pasien
    try:
        send_prescription_ready_email(resep)
    except Exception as e:
        print(f"Error sending email: {e}")
    
    flash('Biaya obat berhasil dikonfirmasi dan tagihan pasien telah diperbarui.', 'success')
    return redirect(url_for('apoteker_dashboard'))

@app.route('/pembayaran')
@login_required
@role_required('pasien')
def lihat_pembayaran():
    # Logika untuk mengambil semua visit milik pasien
    visits = Visit.query.join(Patient).join(User).filter(User.id == current_user.id).all()
    visit_ids = [v.id for v in visits]
    
    # Ambil semua pembayaran yang terkait dengan visit tersebut
    tagihan_list = Pembayaran.query.filter(Pembayaran.visit_id.in_(visit_ids)).order_by(Pembayaran.id.desc()).all()
        
    return render_template('pembayaran.html', tagihan_list=tagihan_list)

@app.route('/manage_pembayaran')
@login_required
@role_required('admin')
def manage_pembayaran():
    tagihan_list = Pembayaran.query.filter(Pembayaran.status != 'Lunas').order_by(Pembayaran.id.desc()).all()
    return render_template('manage_pembayaran.html', tagihan_list=tagihan_list)

@app.route('/pembayaran/<int:pembayaran_id>/konfirmasi', methods=['POST'])
@login_required
@role_required('admin')
def konfirmasi_pembayaran(pembayaran_id):
    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    pembayaran.status = 'Lunas'
    pembayaran.tanggal_lunas = datetime.now(timezone.utc)
    
    resep = ResepObat.query.filter_by(visit_id=pembayaran.visit_id).first()
    if resep:
        resep.status = 'Selesai'
        
    catat_log(f'Admin {current_user.username} mengonfirmasi pembayaran ID: {pembayaran.id}.')
    db.session.commit()
    flash('Pembayaran berhasil dikonfirmasi.', 'success')
    return redirect(url_for('manage_pembayaran'))

# --- TAMBAHKAN 3 FUNGSI BARU INI ---

@app.route('/hubungi_admin', methods=['GET', 'POST'])
@login_required
def hubungi_admin():
    if request.method == 'POST':
        subjek = request.form.get('subjek')
        pesan = request.form.get('pesan')
        
        if not subjek or not pesan:
            flash('Subjek dan Pesan tidak boleh kosong.', 'danger')
            return redirect(url_for('hubungi_admin'))

        new_pesan = PesanSupport(
            user_id=current_user.id,
            subjek=subjek,
            pesan=pesan,
            status='Baru'
        )
        db.session.add(new_pesan)
        db.session.commit()
        
        catat_log(f'User {current_user.username} mengirim pesan support.')
        flash('Pesan Anda telah terkirim ke Admin. Harap tunggu balasan.', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('hubungi_admin.html')

@app.route('/admin/inbox')
@login_required
@role_required('admin')
def admin_inbox():
    pesan_list = PesanSupport.query.order_by(PesanSupport.status.asc(), PesanSupport.timestamp.desc()).all()
    return render_template('admin_inbox.html', pesan_list=pesan_list)

@app.route('/admin/inbox/mark_read/<int:pesan_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_mark_read(pesan_id):
    pesan = PesanSupport.query.get_or_404(pesan_id)
    pesan.status = 'Sudah Dibaca'
    db.session.commit()
    catat_log(f'Admin {current_user.username} menandai pesan ID {pesan.id} sebagai Dibaca.')
    flash('Pesan ditandai sudah dibaca.', 'info')
    return redirect(url_for('admin_inbox'))

# --- Rute Administrasi ---
@app.route('/manage_users')
@login_required
@role_required('admin')
def manage_users():
    unverified_patients = User.query.filter_by(role='pasien', verification_status='belum diverifikasi').all()

    orphaned_patients = User.query.filter(
        User.role == 'pasien',
        User.verification_status == 'aktif',
        User.patient_record == None
    ).all()

    # --- TAMBAHKAN INI ---
    unverified_staff = User.query.filter(
        User.role.in_(['dokter', 'apoteker']),
        User.verification_status == 'menunggu verifikasi data'
    ).all()
    # --- BATAS PENAMBAHAN ---

    all_users = User.query.all()
    return render_template(
        'manage_users.html', 
        users=all_users, 
        unverified_patients=unverified_patients,
        orphaned_patients=orphaned_patients,
        unverified_staff=unverified_staff  # <-- Kirim data baru
    )

@app.route('/user/<int:user_id>/verify', methods=['POST'])
@login_required
@role_required('admin')
def verify_user(user_id):
    user_to_verify = User.query.get_or_404(user_id)

    if user_to_verify.role == 'pasien' and user_to_verify.verification_status == 'belum diverifikasi':
        user_to_verify.verification_status = 'aktif'

        if not user_to_verify.patient_record:
            new_patient_record = Patient(
                user_id=user_to_verify.id,
                medical_record_number=f"RM-{datetime.now().strftime('%Y%m%d%H%M%S')}-{user_to_verify.id}",
                full_name_encrypted=fernet.encrypt(user_to_verify.username.encode()),
                date_of_birth_encrypted=fernet.encrypt(b"1900-01-01"),
                address_encrypted=fernet.encrypt(b"Data Belum Lengkap")
            )
            db.session.add(new_patient_record)

        db.session.commit()
        catat_log(f'Admin {current_user.username} memverifikasi pasien baru: {user_to_verify.username} dan membuat rekam medis.')
        
        # ✨ TAMBAHAN BARU: Kirim email verifikasi
        try:
            send_account_verified_email(user_to_verify)
        except Exception as e:
            print(f"Error sending email: {e}")
        
        flash(f'Pasien "{user_to_verify.username}" telah berhasil diverifikasi.', 'success')

    elif user_to_verify.role in ['dokter', 'apoteker'] and user_to_verify.verification_status == 'menunggu verifikasi data':
        user_to_verify.verification_status = 'aktif'
        db.session.commit()
        catat_log(f'Admin {current_user.username} memverifikasi data profil staf: {user_to_verify.username}.')
        
        # ✨ TAMBAHAN BARU: Kirim email verifikasi
        try:
            send_account_verified_email(user_to_verify)
        except Exception as e:
            print(f"Error sending email: {e}")
        
        flash(f'Data profil untuk "{user_to_verify.username}" telah disetujui.', 'success')

    elif user_to_verify.role == 'admin' and user_to_verify.verification_status == 'belum diverifikasi':
        user_to_verify.verification_status = 'aktif'
        db.session.commit()
        catat_log(f'Admin {current_user.username} mengaktifkan profil admin: {user_to_verify.username}.')
        flash(f'Profil admin "{user_to_verify.username}" telah diaktifkan.', 'success')

    else:
        flash('Aksi verifikasi tidak valid atau tidak diperlukan.', 'warning')

    return redirect(url_for('manage_users'))

@app.route('/user/<int:user_id>/recreate_record', methods=['POST'])
@login_required
@role_required('admin')
def recreate_patient_record(user_id):
    user_to_fix = User.query.get_or_404(user_id)
    
    # Pastikan user adalah pasien dan data 'patient'-nya benar-benar tidak ada
    if user_to_fix.role == 'pasien' and not user_to_fix.patient_record:
        
        # --- PERBAIKAN BUG: Tambahkan data placeholder jika data di user 'None' ---
        # Ini untuk mencegah error jika pasien belum pernah 'complete_profile'
        full_name_data = user_to_fix.full_name_encrypted
        if not full_name_data:
            full_name_data = fernet.encrypt(user_to_fix.username.encode())
            
        dob_data = user_to_fix.date_of_birth_encrypted
        if not dob_data:
            dob_data = fernet.encrypt(b"1900-01-01")
        # --- Batas Perbaikan Bug ---

        # Buat data rekam medis baru
        new_patient_record = Patient(
            user_id=user_to_fix.id,
            medical_record_number=f"RM-{datetime.now().strftime('%Y%m%d%H%M%S')}-{user_to_fix.id}",
            
            # Gunakan data yang sudah divalidasi
            full_name_encrypted=full_name_data,
            date_of_birth_encrypted=dob_data,
            
            # Data alamat diset sebagai placeholder, akan diisi pasien
            address_encrypted=fernet.encrypt(b"Alamat Hilang, Harap Lengkapi Ulang Profil")
        )
        db.session.add(new_patient_record)
        
        # --- FITUR BARU: Paksa pasien melengkapi profil ---
        # Dengan menghapus salah satu data, 'is_profile_complete' akan jadi 'False'
        user_to_fix.contact_encrypted = None 
        # --- Batas Fitur Baru ---
        
        db.session.commit()
        
        catat_log(f'Admin {current_user.username} membuat ulang rekam medis untuk pasien: {user_to_fix.username}.')
        flash(f'Rekam medis untuk "{user_to_fix.username}" telah berhasil dibuat ulang. Pasien akan diminta melengkapi profil saat login.', 'success')
    
    else:
        flash('Aksi tidak valid atau rekam medis pasien sudah ada.', 'warning')
    
    return redirect(url_for('manage_users'))

@app.route('/create_user', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def create_user_by_admin():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        email = request.form.get('email')

        if role not in ['admin', 'dokter', 'apoteker']:
            flash('Peran tidak valid.', 'danger')
            return redirect(url_for('create_user_by_admin'))

        if User.query.filter_by(username=username).first():
            flash('Username sudah digunakan.', 'danger')
            return redirect(url_for('create_user_by_admin'))
        
        if not email:
            flash('Email wajib diisi.', 'danger')
            return redirect(url_for('create_user_by_admin'))
        if User.query.filter_by(email=email).first():
            flash('Email sudah digunakan.', 'danger')
            return redirect(url_for('create_user_by_admin'))

        new_user = User(username=username, role=role, email=email, verification_status='belum diverifikasi')
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        catat_log(f'Admin {current_user.username} membuat user baru: {username} ({role}).')
        flash(f'Akun "{username}" berhasil dibuat.', 'success')
        return redirect(url_for('manage_users'))
    return render_template('create_user.html')

@app.route('/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_user(user_id):
    user_to_edit = User.query.get_or_404(user_id)
    if request.method == 'POST':
        new_email = request.form.get('email')
        user_to_edit.email = new_email
        user_to_edit.username = request.form.get('username')
        user_to_edit.role = request.form.get('role')
        password = request.form.get('password')
        
        if not new_email:
            flash('Email wajib diisi.', 'danger')
            return render_template('edit_user.html', user=user_to_edit)

        existing_user = User.query.filter(User.email == new_email, User.id != user_id).first()
        if existing_user:
            flash('Email tersebut sudah digunakan oleh pengguna lain.', 'danger')
            return render_template('edit_user.html', user=user_to_edit)
        
        if password:
            user_to_edit.set_password(password)
        db.session.commit()
        catat_log(f'Admin {current_user.username} mengubah data user: {user_to_edit.username}.')
        flash('Data pengguna berhasil diperbarui!', 'success')
        return redirect(url_for('manage_users'))
    return render_template('edit_user.html', user=user_to_edit)

@app.route('/user/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def delete_user(user_id):
    if current_user.id == user_id:
        flash('Anda tidak dapat menonaktifkan akun Anda sendiri.', 'warning')
        return redirect(url_for('manage_users'))

    user_to_deactivate = User.query.get_or_404(user_id)
    username = user_to_deactivate.username

    try:
        # --- UBAH INI: Set is_active = False ---
        user_to_deactivate.is_active_db = False
        db.session.commit()
        # --- BATAS PERUBAHAN ---

        catat_log(f'Admin {current_user.username} menonaktifkan user: {username}.')
        flash(f'Pengguna "{username}" berhasil dinonaktifkan.', 'success')

    except Exception as e:
        db.session.rollback() # Batalkan jika ada error
        flash(f'Gagal menonaktifkan pengguna: {e}', 'danger')

    return redirect(url_for('manage_users'))

@app.route('/schedule/<string:request_id>/request_finish', methods=['POST'])
@login_required
@role_required('dokter')
def request_finish_schedule(request_id):
    # Ambil semua slot yang statusnya 'Approved' di grup ini
    slots_to_request_finish = AppointmentSlot.query.filter_by(
        schedule_request_id=request_id, 
        doctor_id=current_user.id,
        verification_status='Approved' # Hanya bisa ajukan yang statusnya Approved
    ).all()

    if not slots_to_request_finish:
        flash('Jadwal tidak ditemukan atau statusnya tidak valid.', 'danger')
        return redirect(url_for('doctor_schedule'))
    
    try:
        for slot in slots_to_request_finish:
            # Ubah status semua slot di grup ini menjadi 'Menunggu Penyelesaian'
            slot.verification_status = 'Menunggu Penyelesaian'
        
        db.session.commit()
        catat_log(f'Dokter {current_user.username} mengajukan penyelesaian jadwal ID {request_id}.')
        flash('Pengajuan untuk menyelesaikan jadwal telah dikirim ke Admin.', 'info')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi error saat mengajukan penyelesaian: {e}', 'danger')

    return redirect(url_for('doctor_schedule'))

@app.route('/user/<int:user_id>/undelete', methods=['POST'])
@login_required
@role_required('admin')
def undelete_user(user_id):
    user_to_reactivate = User.query.get_or_404(user_id)
    username = user_to_reactivate.username

    try:
        user_to_reactivate.is_active_db = True
        db.session.commit()
        catat_log(f'Admin {current_user.username} mengaktifkan kembali user: {username}.')
        flash(f'Pengguna "{username}" berhasil diaktifkan kembali.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal mengaktifkan kembali pengguna: {e}', 'danger')

    return redirect(url_for('manage_users'))

@app.route('/reports')
@login_required
@role_required('admin', 'dokter')
@verified_staff_required
def reports():
    # 1. Statistik (Biarkan seperti semula)
    patient_count = Patient.query.count()
    doctor_count = User.query.filter_by(role='dokter').count()
    visit_count = Visit.query.count()

    # 2. Ambil parameter filter dari URL
    search_rm = request.args.get('q_rm', '')
    search_doctor = request.args.get('q_doctor', '')
    search_start = request.args.get('date_start', '')
    search_end = request.args.get('date_end', '')

    # 3. Query dasar untuk mengambil kunjungan
    # Kita join semua tabel yang mungkin diperlukan untuk filter
    visit_query = db.session.query(Visit).join(
        Patient, Visit.patient_id == Patient.id
    ).join(
        User, Visit.doctor_id == User.id
    ).outerjoin(
        SoapNote, Visit.id == SoapNote.visit_id
    )

    # 4. Terapkan filter secara dinamis
    if search_rm:
        visit_query = visit_query.filter(Patient.medical_record_number.like(f'%{search_rm}%'))
    
    if search_doctor:
        visit_query = visit_query.filter(User.username.like(f'%{search_doctor}%'))
        
    if search_start:
        try:
            start_date = datetime.strptime(search_start, '%Y-%m-%d')
            visit_query = visit_query.filter(Visit.visit_date >= start_date)
        except ValueError:
            flash('Format Tanggal Mulai tidak valid. Gunakan YYYY-MM-DD.', 'warning')
    
    if search_end:
        try:
            # Tambah 1 hari agar tanggal akhir bersifat inklusif (mencakup data di hari tsb)
            end_date = datetime.strptime(search_end, '%Y-%m-%d') + timedelta(days=1)
            visit_query = visit_query.filter(Visit.visit_date < end_date)
        except ValueError:
            flash('Format Tanggal Selesai tidak valid. Gunakan YYYY-MM-DD.', 'warning')

    # 5. Eksekusi query
    visits = visit_query.order_by(Visit.visit_date.desc()).all()

    return render_template(
        'reports.html',
        patient_count=patient_count,
        doctor_count=doctor_count,
        visit_count=visit_count,
        visits=visits,  # Ganti 'recent_visits' menjadi 'visits'
        
        # Kirim kembali nilai filter ke template
        search_rm=search_rm,
        search_doctor=search_doctor,
        search_start=search_start,
        search_end=search_end
    )

@app.route('/patient/<int:patient_id>/export_pdf')
@login_required
@role_required('admin', 'dokter')
def export_patient_to_pdf(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    
    # Buat buffer untuk PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           leftMargin=2*cm, rightMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    # Container untuk elemen PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Style custom
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#4781ff'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    # Header
    elements.append(Paragraph("REKAM MEDIS PASIEN", title_style))
    elements.append(Paragraph("Batam Sehat - Sistem Informasi Kesehatan", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Data Pasien
    patient_data = [
        ['No. Rekam Medis:', patient.medical_record_number],
        ['Nama Lengkap:', patient.full_name],
        ['Tanggal Lahir:', patient.date_of_birth.strftime('%d %B %Y') if patient.date_of_birth else '-'],
        ['Usia:', f"{patient.age} tahun" if patient.age else '-'],
        ['Alamat:', patient.address],
    ]
    
    if patient.user:
        patient_data.extend([
            ['Kontak:', patient.user.contact],
            ['Email:', patient.user.email or '-']
        ])
    
    patient_table = Table(patient_data, colWidths=[5*cm, 12*cm])
    patient_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    
    elements.append(patient_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Riwayat Kunjungan
    elements.append(Paragraph("RIWAYAT KUNJUNGAN", styles['Heading2']))
    elements.append(Spacer(1, 0.3*cm))
    
    visits = Visit.query.filter_by(patient_id=patient.id).order_by(Visit.visit_date.desc()).all()
    
    for idx, visit in enumerate(visits, 1):
        visit_title = f"Kunjungan #{idx} - {visit.visit_date.strftime('%d %B %Y, %H:%M')}"
        elements.append(Paragraph(visit_title, styles['Heading3']))
        
        visit_data = [
            ['Dokter:', visit.doctor.username],
            ['Subjective:', visit.soap_note.subjective or '-'],
            ['Objective:', visit.soap_note.objective or '-'],
            ['Assessment:', visit.soap_note.assessment or '-'],
            ['Plan:', visit.soap_note.plan or '-']
        ]
        
        if visit.resep_obat:
            visit_data.append(['Resep Obat:', visit.resep_obat.detail_resep])
        
        visit_table = Table(visit_data, colWidths=[4*cm, 13*cm])
        visit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4ff')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(visit_table)
        elements.append(Spacer(1, 0.4*cm))
    
    if not visits:
        elements.append(Paragraph("Belum ada riwayat kunjungan.", styles['Normal']))
    
    # Footer
    elements.append(Spacer(1, 1*cm))
    footer_text = f"Dicetak oleh: {current_user.username} | Tanggal: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    catat_log(f'{current_user.role.capitalize()} {current_user.username} mengexport rekam medis pasien {patient.full_name} ke PDF.')
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Rekam_Medis_{patient.medical_record_number}.pdf'
    )

@app.route('/pembayaran/<int:pembayaran_id>/export_pdf')
@login_required
def export_invoice_to_pdf(pembayaran_id):
    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    
    # Validasi akses
    if current_user.role == 'pasien':
        if pembayaran.visit.patient.user_id != current_user.id:
            flash('Anda tidak memiliki akses ke invoice ini.', 'danger')
            return redirect(url_for('lihat_pembayaran'))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Header Invoice
    title_style = ParagraphStyle('InvoiceTitle', parent=styles['Heading1'],
                                 fontSize=18, textColor=colors.HexColor('#4781ff'),
                                 alignment=1)
    elements.append(Paragraph("INVOICE PEMBAYARAN", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Info Invoice
    invoice_info = [
        ['No. Invoice:', f'INV-{pembayaran.id:05d}'],
        ['Tanggal Kunjungan:', pembayaran.visit.visit_date.strftime('%d %B %Y')],
        ['Status Pembayaran:', pembayaran.status]
    ]
    
    if pembayaran.tanggal_lunas:
        invoice_info.append(['Tanggal Lunas:', pembayaran.tanggal_lunas.strftime('%d %B %Y, %H:%M')])
    
    info_table = Table(invoice_info, colWidths=[5*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Data Pasien
    patient = pembayaran.visit.patient
    elements.append(Paragraph("INFORMASI PASIEN", styles['Heading3']))
    patient_info = [
        ['Nama:', patient.full_name],
        ['No. RM:', patient.medical_record_number],
        ['Dokter:', pembayaran.visit.doctor.username]
    ]
    patient_table = Table(patient_info, colWidths=[5*cm, 12*cm])
    patient_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(patient_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Rincian Biaya
    elements.append(Paragraph("RINCIAN BIAYA", styles['Heading3']))
    
    biaya_data = [
        ['Deskripsi', 'Jumlah (Rp)'],
        ['Biaya Konsultasi', f"{pembayaran.biaya_konsultasi:,.0f}"],
        ['Biaya Obat', f"{pembayaran.biaya_obat:,.0f}"],
        ['', ''],
        ['TOTAL', f"{pembayaran.jumlah:,.0f}"]
    ]
    
    biaya_table = Table(biaya_data, colWidths=[12*cm, 5*cm])
    biaya_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4781ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#4781ff'))
    ]))
    elements.append(biaya_table)
    
    # Footer
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph("Terima kasih atas kepercayaan Anda.", styles['Normal']))
    elements.append(Paragraph(f"Dicetak: {datetime.now().strftime('%d %B %Y, %H:%M')}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    catat_log(f'User {current_user.username} mengexport invoice ID {pembayaran.id} ke PDF.')
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Invoice_{pembayaran.id}.pdf'
    )
    
@app.route('/reports/export_excel')
@login_required
@role_required('admin', 'dokter')
def export_reports_to_excel():
    # Ambil filter dari query string (sama seperti halaman reports)
    search_rm = request.args.get('q_rm', '')
    search_doctor = request.args.get('q_doctor', '')
    search_start = request.args.get('date_start', '')
    search_end = request.args.get('date_end', '')
    
    # Query data (sama seperti di halaman reports)
    visit_query = db.session.query(Visit).join(
        Patient, Visit.patient_id == Patient.id
    ).join(
        User, Visit.doctor_id == User.id
    )
    
    if search_rm:
        visit_query = visit_query.filter(Patient.medical_record_number.like(f'%{search_rm}%'))
    
    if search_doctor:
        visit_query = visit_query.filter(User.username.like(f'%{search_doctor}%'))
    
    if search_start:
        try:
            start_date = datetime.strptime(search_start, '%Y-%m-%d')
            visit_query = visit_query.filter(Visit.visit_date >= start_date)
        except ValueError:
            pass
    
    if search_end:
        try:
            end_date = datetime.strptime(search_end, '%Y-%m-%d') + timedelta(days=1)
            visit_query = visit_query.filter(Visit.visit_date < end_date)
        except ValueError:
            pass
    
    visits = visit_query.order_by(Visit.visit_date.desc()).all()
    
    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Kunjungan"
    
    # Header styling
    header_fill = PatternFill(start_color='4781FF', end_color='4781FF', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Header columns
    headers = ['Tanggal Kunjungan', 'No. RM', 'Nama Pasien', 'Usia', 'Dokter', 
               'Subjective', 'Objective', 'Assessment', 'Plan', 'Resep Obat']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, visit in enumerate(visits, 2):
        ws.cell(row=row_num, column=1, value=visit.visit_date.strftime('%d-%m-%Y %H:%M'))
        ws.cell(row=row_num, column=2, value=visit.patient.medical_record_number)
        ws.cell(row=row_num, column=3, value=visit.patient.full_name)
        ws.cell(row=row_num, column=4, value=visit.patient.age or '-')
        ws.cell(row=row_num, column=5, value=visit.doctor.username)
        ws.cell(row=row_num, column=6, value=visit.soap_note.subjective or '-')
        ws.cell(row=row_num, column=7, value=visit.soap_note.objective or '-')
        ws.cell(row=row_num, column=8, value=visit.soap_note.assessment or '-')
        ws.cell(row=row_num, column=9, value=visit.soap_note.plan or '-')
        ws.cell(row=row_num, column=10, value=visit.resep_obat.detail_resep if visit.resep_obat else '-')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    catat_log(f'{current_user.role.capitalize()} {current_user.username} mengexport laporan kunjungan ke Excel.')
    
    filename = f'Laporan_Kunjungan_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/analytics/visits_trend')
@login_required
@role_required('admin', 'dokter')
@limiter.limit("100 per minute")
def api_visits_trend():
    """API untuk grafik tren kunjungan 7 hari terakhir"""
    seven_days_ago = datetime.now() - timedelta(days=7)
    
    visits = db.session.query(
        func.date(Visit.visit_date).label('date'),
        func.count(Visit.id).label('count')
    ).filter(
        Visit.visit_date >= seven_days_ago
    ).group_by(
        func.date(Visit.visit_date)
    ).order_by('date').all()
    
    # Format data untuk Chart.js
    labels = []
    data = []
    
    for i in range(7):
        date = (datetime.now() - timedelta(days=6-i)).date()
        labels.append(date.strftime('%d %b'))
        
        # Cari data untuk tanggal ini
        count = next((v.count for v in visits if v.date == date), 0)
        data.append(count)
    
    return jsonify({
        'labels': labels,
        'datasets': [{
            'label': 'Jumlah Kunjungan',
            'data': data,
            'borderColor': 'rgb(71, 129, 255)',
            'backgroundColor': 'rgba(71, 129, 255, 0.1)',
            'tension': 0.3
        }]
    })


@app.route('/api/analytics/revenue_monthly')
@login_required
@role_required('admin')
@limiter.limit("100 per minute")
def api_revenue_monthly():
    """API untuk grafik pendapatan bulanan (6 bulan terakhir)"""
    six_months_ago = datetime.now() - timedelta(days=180)
    
    payments = db.session.query(
        extract('year', Visit.visit_date).label('year'),
        extract('month', Visit.visit_date).label('month'),
        func.sum(Pembayaran.biaya_konsultasi + Pembayaran.biaya_obat).label('total')
    ).join(
        Visit, Pembayaran.visit_id == Visit.id
    ).filter(
        Visit.visit_date >= six_months_ago,
        Pembayaran.status == 'Lunas'
    ).group_by('year', 'month').order_by('year', 'month').all()
    
    # Format data
    labels = []
    data = []
    
    for i in range(6):
        date = datetime.now() - timedelta(days=30 * (5 - i))
        month_name = date.strftime('%b %Y')
        labels.append(month_name)
        
        # Cari total untuk bulan ini
        total = next((p.total for p in payments if p.year == date.year and p.month == date.month), 0)
        data.append(float(total) if total else 0)
    
    return jsonify({
        'labels': labels,
        'datasets': [{
            'label': 'Pendapatan (Rp)',
            'data': data,
            'backgroundColor': 'rgba(40, 167, 69, 0.5)',
            'borderColor': 'rgb(40, 167, 69)',
            'borderWidth': 2
        }]
    })


@app.route('/api/analytics/top_doctors')
@login_required
@role_required('admin')
@limiter.limit("100 per minute")
def api_top_doctors():
    """API untuk grafik dokter dengan kunjungan terbanyak"""
    top_doctors = db.session.query(
        User.username,
        func.count(Visit.id).label('visit_count')
    ).join(
        Visit, User.id == Visit.doctor_id
    ).filter(
        User.role == 'dokter'
    ).group_by(User.id).order_by(func.count(Visit.id).desc()).limit(5).all()
    
    labels = [doc.username for doc in top_doctors]
    data = [doc.visit_count for doc in top_doctors]
    
    return jsonify({
        'labels': labels,
        'datasets': [{
            'label': 'Jumlah Kunjungan',
            'data': data,
            'backgroundColor': [
                'rgba(255, 99, 132, 0.6)',
                'rgba(54, 162, 235, 0.6)',
                'rgba(255, 206, 86, 0.6)',
                'rgba(75, 192, 192, 0.6)',
                'rgba(153, 102, 255, 0.6)'
            ]
        }]
    })


@app.route('/api/analytics/payment_status')
@login_required
@role_required('admin')
@limiter.limit("100 per minute")
def api_payment_status():
    """API untuk grafik status pembayaran (Pie Chart)"""
    payments = db.session.query(
        Pembayaran.status,
        func.count(Pembayaran.id).label('count')
    ).group_by(Pembayaran.status).all()
    
    labels = [p.status for p in payments]
    data = [p.count for p in payments]
    
    return jsonify({
        'labels': labels,
        'datasets': [{
            'data': data,
            'backgroundColor': [
                'rgba(40, 167, 69, 0.7)',   # Lunas - hijau
                'rgba(220, 53, 69, 0.7)',   # Belum Lunas - merah
                'rgba(255, 193, 7, 0.7)'    # Pending - kuning
            ]
        }]
    })

@app.route('/admin/analytics')
@login_required
@role_required('admin')
def admin_analytics():
    """Halaman analytics lengkap untuk admin"""
    
    # Statistik umum
    total_patients = Patient.query.count()
    total_doctors = User.query.filter_by(role='dokter').count()
    total_visits = Visit.query.count()
    
    # Pendapatan bulan ini
    this_month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_revenue = db.session.query(
        func.sum(Pembayaran.biaya_konsultasi + Pembayaran.biaya_obat)
    ).join(Visit).filter(
        Visit.visit_date >= this_month_start,
        Pembayaran.status == 'Lunas'
    ).scalar() or 0
    
    # Appointment pending
    pending_appointments = Appointment.query.filter_by(status='Pending').count()
    
    # Tagihan belum lunas
    unpaid_bills = Pembayaran.query.filter_by(status='Belum Lunas').count()
    
    return render_template('admin_analytics.html',
                         total_patients=total_patients,
                         total_doctors=total_doctors,
                         total_visits=total_visits,
                         this_month_revenue=this_month_revenue,
                         pending_appointments=pending_appointments,
                         unpaid_bills=unpaid_bills)

@app.route('/dashboard_unverified')
@login_required
@role_required('dokter', 'apoteker')
def dashboard_unverified_staff():
    """Dashboard terbatas untuk staf yang belum terverifikasi"""
    if current_user.verification_status == 'aktif':
        return redirect(url_for('dashboard'))
    
    return render_template('dashboard_unverified_staff.html')

@app.route('/api/analytics/my_visits_trend')
@login_required
@role_required('dokter')
@limiter.limit("100 per minute")
def api_my_visits_trend():
    """API untuk grafik tren kunjungan dokter sendiri"""
    seven_days_ago = datetime.now() - timedelta(days=7)
    
    visits = db.session.query(
        func.date(Visit.visit_date).label('date'),
        func.count(Visit.id).label('count')
    ).filter(
        Visit.doctor_id == current_user.id,
        Visit.visit_date >= seven_days_ago
    ).group_by(
        func.date(Visit.visit_date)
    ).order_by('date').all()
    
    labels = []
    data = []
    
    for i in range(7):
        date = (datetime.now() - timedelta(days=6-i)).date()
        labels.append(date.strftime('%d %b'))
        count = next((v.count for v in visits if v.date == date), 0)
        data.append(count)
    
    return jsonify({
        'labels': labels,
        'datasets': [{
            'label': 'Pasien Saya',
            'data': data,
            'borderColor': 'rgb(71, 129, 255)',
            'backgroundColor': 'rgba(71, 129, 255, 0.1)',
            'tension': 0.3
        }]
    })

@app.route('/staff/login', methods=['GET', 'POST'])
@limiter.limit("10 per 5 minutes")
def staff_login():
    """Login khusus untuk Staff (Admin, Dokter, Apoteker)"""
    if current_user.is_authenticated:
        if current_user.role in ['admin', 'dokter', 'apoteker']:
            return redirect(url_for('dashboard'))
        else:
            logout_user()
            flash('Silakan gunakan login pasien.', 'warning')
            return redirect(url_for('patient_login'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Username dan password wajib diisi.', 'danger')
            return render_template('staff_login.html')
        
        can_login, error_msg = check_login_attempts(username, 'staff')
        if not can_login:
            flash(error_msg, 'danger')
            catat_log(f'Login staff gagal (terkunci): {username}')
            return render_template('staff_login.html')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.role not in ['admin', 'dokter', 'apoteker']:
            flash('Akses ditolak. Gunakan login pasien untuk akun Anda.', 'danger')
            catat_log(f'Percobaan login staff dengan akun pasien: {username}')
            return render_template('staff_login.html')
        
        if user and user.check_password(password):
            if not user.is_active_db:
                flash('Akun Anda tidak aktif. Hubungi administrator.', 'danger')
                catat_log(f'Login staff gagal (akun nonaktif): {username}')
                return render_template('staff_login.html')
            
            reset_login_attempts(username, 'staff')
            
            session['session_token'] = generate_session_token()
            session['session_ip'] = hash_ip(request.remote_addr)
            session['login_time'] = datetime.now().isoformat()
            session.permanent = True
            
            if user.otp_secret:
                session['user_id_2fa'] = user.id
                session['2fa_role'] = 'staff'
                return redirect(url_for('staff_login_2fa'))
            else:
                login_user(user)
                catat_log(f'Staff {username} ({user.role}) berhasil login (tanpa 2FA).')
                flash(f'Selamat datang, {user.username}!', 'success')
                return redirect(url_for('dashboard'))
        else:
            is_locked = record_failed_login(username, 'staff')
            
            if is_locked:
                flash('Terlalu banyak percobaan gagal. Akun terkunci selama 20 menit.', 'danger')
                catat_log(f'Login staff gagal (akun terkunci): {username}')
            else:
                remaining = 8 - session.get(f"login_attempts_staff_{username}", 0)
                flash(f'Login gagal. Username atau password salah. ({remaining} percobaan tersisa sebelum akun terkunci)', 'danger')
                catat_log(f'Login staff gagal: {username}')
            
            return render_template('staff_login.html')
    
    return render_template('staff_login.html')

@app.route('/staff/login/2fa', methods=['GET', 'POST'])
def staff_login_2fa():
    """2FA untuk staff login"""
    if 'user_id_2fa' not in session or session.get('2fa_role') != 'staff':
        return redirect(url_for('staff_login'))

    user_id = session['user_id_2fa']
    user = User.query.get(user_id)
    
    if not user or user.role not in ['admin', 'dokter', 'apoteker']:
        session.pop('user_id_2fa', None)
        session.pop('2fa_role', None)
        return redirect(url_for('staff_login'))

    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        
        if not token or len(token) != 6:
            flash('Kode OTP tidak valid.', 'danger')
            return render_template('staff_login_2fa.html')
        
        totp = pyotp.TOTP(user.otp_secret)
        if totp.verify(token):
            session.pop('user_id_2fa', None)
            session.pop('2fa_role', None)
            
            # Set session security
            session['session_token'] = generate_session_token()
            session['session_ip'] = hash_ip(request.remote_addr)
            session['login_time'] = datetime.now().isoformat()
            
            login_user(user)
            reset_login_attempts(user.username, 'staff')
            catat_log(f'Staff {user.username} ({user.role}) berhasil login (dengan 2FA).')
            flash(f'Selamat datang, {user.username}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Kode OTP tidak valid.', 'danger')
            catat_log(f'2FA staff gagal: {user.username}')
    
    return render_template('staff_login_2fa.html')

@app.route('/patient/login', methods=['GET', 'POST'])
@limiter.limit("10 per 5 minutes")
def patient_login():
    """Login khusus untuk Pasien"""
    if current_user.is_authenticated:
        if current_user.role == 'pasien':
            return redirect(url_for('dashboard'))
        else:
            logout_user()
            flash('Silakan gunakan login staff.', 'warning')
            return redirect(url_for('staff_login'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Username dan password wajib diisi.', 'danger')
            return render_template('patient_login.html')
        
        can_login, error_msg = check_login_attempts(username, 'patient')
        if not can_login:
            flash(error_msg, 'danger')
            catat_log(f'Login pasien gagal (terkunci): {username}')
            return render_template('patient_login.html')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.role != 'pasien':
            flash('Akses ditolak. Gunakan login staff untuk akun Anda.', 'danger')
            catat_log(f'Percobaan login pasien dengan akun staff: {username}')
            return render_template('patient_login.html')
        
        if user and user.check_password(password):
            if not user.is_active_db:
                flash('Akun Anda tidak aktif. Hubungi administrator.', 'danger')
                catat_log(f'Login pasien gagal (akun nonaktif): {username}')
                return render_template('patient_login.html')
            
            reset_login_attempts(username, 'patient')
            
            session['session_token'] = generate_session_token()
            session['session_ip'] = hash_ip(request.remote_addr)
            session['login_time'] = datetime.now().isoformat()
            session.permanent = True
            
            if user.otp_secret:
                session['user_id_2fa'] = user.id
                session['2fa_role'] = 'patient'
                return redirect(url_for('patient_login_2fa'))
            else:
                login_user(user)
                catat_log(f'Pasien {username} berhasil login (tanpa 2FA).')
                flash(f'Selamat datang, {user.username}!', 'success')
                return redirect(url_for('dashboard'))
        else:
            is_locked = record_failed_login(username, 'patient')
            
            if is_locked:
                flash('Terlalu banyak percobaan gagal. Akun terkunci selama 20 menit.', 'danger')
                catat_log(f'Login pasien gagal (akun terkunci): {username}')
            else:
                remaining = 8 - session.get(f"login_attempts_patient_{username}", 0)
                flash(f'Login gagal. Username atau password salah. ({remaining} percobaan tersisa sebelum akun terkunci)', 'danger')
                catat_log(f'Login pasien gagal: {username}')
            
            return render_template('patient_login.html')
    
    return render_template('patient_login.html')

@app.route('/patient/login/2fa', methods=['GET', 'POST'])
def patient_login_2fa():
    """2FA untuk patient login"""
    if 'user_id_2fa' not in session or session.get('2fa_role') != 'patient':
        return redirect(url_for('patient_login'))

    user_id = session['user_id_2fa']
    user = User.query.get(user_id)
    
    if not user or user.role != 'pasien':
        session.pop('user_id_2fa', None)
        session.pop('2fa_role', None)
        return redirect(url_for('patient_login'))

    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        
        if not token or len(token) != 6:
            flash('Kode OTP tidak valid.', 'danger')
            return render_template('patient_login_2fa.html')
        
        totp = pyotp.TOTP(user.otp_secret)
        if totp.verify(token):
            session.pop('user_id_2fa', None)
            session.pop('2fa_role', None)
            
            # Set session security
            session['session_token'] = generate_session_token()
            session['session_ip'] = hash_ip(request.remote_addr)
            session['login_time'] = datetime.now().isoformat()
            
            login_user(user)
            reset_login_attempts(user.username, 'patient')
            catat_log(f'Pasien {user.username} berhasil login (dengan 2FA).')
            flash(f'Selamat datang, {user.username}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Kode OTP tidak valid.', 'danger')
            catat_log(f'2FA pasien gagal: {user.username}')
    
    return render_template('patient_login_2fa.html')

@app.route('/toggle_theme', methods=['POST'])
def toggle_theme():
    """Toggle antara light dan dark mode"""
    current_theme = session.get('theme', 'light')
    new_theme = 'dark' if current_theme == 'light' else 'light'
    session['theme'] = new_theme
    session.permanent = True
    
    if request.is_json:
        return jsonify({'success': True, 'theme': new_theme})
    
    return redirect(request.referrer or url_for('dashboard'))

@app.context_processor
def inject_theme():
    """Inject theme variable ke semua template"""
    return dict(current_theme=session.get('theme', 'light'))

@app.errorhandler(429)
def ratelimit_handler(e):
    """Handler untuk error rate limit exceeded"""
    return render_template('error_429.html', 
                         description=e.description,
                         retry_after=e.description.split()[-2] if 'in' in e.description else 'beberapa saat'), 429

if __name__ == '__main__':
    app.run(debug=True)